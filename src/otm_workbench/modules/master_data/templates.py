import json
import zipfile
from datetime import UTC, datetime
from pathlib import Path
from typing import BinaryIO

from openpyxl import load_workbook
from openpyxl import Workbook
from sqlalchemy.orm import Session

from otm_workbench.catalog.services import validate_column, validate_table
from otm_workbench.models import (
    Artifact,
    AuditLog,
    Evidence,
    Manifest,
    MasterDataBatch,
    MasterDataCanonicalRecord,
    MasterDataCsvFile,
    MasterDataOutputRecord,
    MasterDataTemplate,
)
from otm_workbench.modules.rates.csv_preview import build_otm_csv_preview
from otm_workbench.platform.services import file_sha256


REGIONS_BASIC_TEMPLATE = {
    "code": "REGIONS_BASIC",
    "name": "Regions Basic",
    "version": 1,
    "status": "PUBLISHED",
    "catalog_macro_object_code": "REGION",
    "data_category": "MASTER_DATA",
    "target_tables": ["REGION", "REGION_DETAIL"],
    "sheets": [
        {
            "code": "REGIONS",
            "name": "Regions",
            "target_table": "REGION",
            "fields": [
                {
                    "name": "region_gid",
                    "label": "Region GID",
                    "target_column": "REGION_GID",
                    "required": True,
                    "data_type": "string",
                },
                {
                    "name": "region_xid",
                    "label": "Region XID",
                    "target_column": "REGION_XID",
                    "required": True,
                    "data_type": "string",
                },
                {
                    "name": "region_name",
                    "label": "Region Name",
                    "target_column": "REGION_NAME",
                    "required": False,
                    "data_type": "string",
                },
            ],
        },
        {
            "code": "REGION_DETAILS",
            "name": "Region Details",
            "target_table": "REGION_DETAIL",
            "fields": [
                {
                    "name": "region_gid",
                    "label": "Region GID",
                    "target_column": "REGION_GID",
                    "required": True,
                    "data_type": "string",
                },
                {
                    "name": "location_gid",
                    "label": "Location GID",
                    "target_column": "LOCATION_GID",
                    "required": True,
                    "data_type": "string",
                },
            ],
        },
    ],
    "description": "Synthetic starter template for region master data.",
}

ITEMS_PACKAGING_STANDARD_TEMPLATE = {
    "code": "ITEMS_PACKAGING_STANDARD",
    "name": "Items & Packaging Standard",
    "version": 1,
    "status": "PUBLISHED",
    "catalog_macro_object_code": "ITEM",
    "data_category": "MASTER_DATA",
    "target_tables": ["ITEM", "SHIP_UNIT_SPEC", "PACKAGED_ITEM", "TI_HI"],
    "sheets": [
        {
            "code": "ITEMS",
            "name": "Items",
            "target_table": "ITEM",
            "fields": [
                {
                    "name": "item_gid",
                    "label": "Item GID",
                    "target_column": "ITEM_GID",
                    "required": True,
                    "data_type": "string",
                },
                {
                    "name": "item_xid",
                    "label": "Item XID",
                    "target_column": "ITEM_XID",
                    "required": True,
                    "data_type": "string",
                },
                {
                    "name": "description",
                    "label": "Description",
                    "target_column": "DESCRIPTION",
                    "required": False,
                    "data_type": "string",
                },
                {
                    "name": "item_type_gid",
                    "label": "Item Type GID",
                    "target_column": "ITEM_TYPE_GID",
                    "required": False,
                    "data_type": "string",
                },
            ],
        },
        {
            "code": "PACKAGING",
            "name": "Packaging",
            "target_table": "PACKAGED_ITEM",
            "fields": [
                {
                    "name": "packaged_item_gid",
                    "label": "Packaged Item GID",
                    "target_column": "PACKAGED_ITEM_GID",
                    "required": True,
                    "data_type": "string",
                },
                {
                    "name": "packaged_item_xid",
                    "label": "Packaged Item XID",
                    "target_column": "PACKAGED_ITEM_XID",
                    "required": True,
                    "data_type": "string",
                },
                {
                    "name": "item_gid",
                    "label": "Item GID",
                    "target_column": "ITEM_GID",
                    "required": True,
                    "data_type": "string",
                },
                {
                    "name": "packaging_unit_gid",
                    "label": "Packaging Unit GID",
                    "target_column": "PACKAGING_UNIT_GID",
                    "required": False,
                    "data_type": "string",
                },
                {
                    "name": "weight",
                    "label": "Weight",
                    "target_column": "PACKAGE_SHIP_UNIT_WEIGHT",
                    "required": False,
                    "data_type": "number",
                },
                {
                    "name": "weight_uom",
                    "label": "Weight UOM",
                    "target_column": "PACKAGE_SHIP_UNIT_WEIGHT_UOM",
                    "required": False,
                    "data_type": "string",
                },
                {
                    "name": "volume",
                    "label": "Volume",
                    "target_column": "PACKAGE_SU_VOLUME",
                    "required": False,
                    "data_type": "number",
                },
                {
                    "name": "volume_uom",
                    "label": "Volume UOM",
                    "target_column": "PACKAGE_SU_VOLUME_UOM_CODE",
                    "required": False,
                    "data_type": "string",
                },
                {
                    "name": "length",
                    "label": "Length",
                    "target_column": "PACKAGE_SU_LENGTH",
                    "required": False,
                    "data_type": "number",
                },
                {
                    "name": "width",
                    "label": "Width",
                    "target_column": "PACKAGE_SU_WIDTH",
                    "required": False,
                    "data_type": "number",
                },
            ],
        },
        {
            "code": "TI_HI",
            "name": "TI HI",
            "target_table": "TI_HI",
            "fields": [
                {
                    "name": "packaged_item_gid",
                    "label": "Packaged Item GID",
                    "target_column": "PACKAGED_ITEM_GID",
                    "required": True,
                    "data_type": "string",
                },
                {
                    "name": "packaging_unit_gid",
                    "label": "Packaging Unit GID",
                    "target_column": "PACKAGING_UNIT_GID",
                    "required": False,
                    "data_type": "string",
                },
                {
                    "name": "num_layers",
                    "label": "Number of Layers",
                    "target_column": "NUM_LAYERS",
                    "required": False,
                    "data_type": "number",
                },
                {
                    "name": "quantity_per_layer",
                    "label": "Quantity per Layer",
                    "target_column": "QUANTITY_PER_LAYER",
                    "required": False,
                    "data_type": "number",
                },
            ],
        },
    ],
    "description": "Synthetic starter template for item and packaged item master data.",
}

LOCATIONS_BASIC_TEMPLATE = {
    "code": "LOCATIONS_BASIC",
    "name": "Locations Basic",
    "version": 1,
    "status": "PUBLISHED",
    "catalog_macro_object_code": "LOCATION",
    "data_category": "MASTER_DATA",
    "target_tables": ["LOCATION", "LOCATION_ADDRESS"],
    "sheets": [
        {
            "code": "LOCATIONS",
            "name": "Locations",
            "target_table": "LOCATION",
            "fields": [
                {
                    "name": "location_gid",
                    "label": "Location GID",
                    "target_column": "LOCATION_GID",
                    "required": True,
                    "data_type": "string",
                },
                {
                    "name": "location_xid",
                    "label": "Location XID",
                    "target_column": "LOCATION_XID",
                    "required": True,
                    "data_type": "string",
                },
                {
                    "name": "location_name",
                    "label": "Location Name",
                    "target_column": "LOCATION_NAME",
                    "required": False,
                    "data_type": "string",
                },
                {
                    "name": "city",
                    "label": "City",
                    "target_column": "CITY",
                    "required": False,
                    "data_type": "string",
                },
                {
                    "name": "province_code",
                    "label": "Province Code",
                    "target_column": "PROVINCE_CODE",
                    "required": False,
                    "data_type": "string",
                },
                {
                    "name": "postal_code",
                    "label": "Postal Code",
                    "target_column": "POSTAL_CODE",
                    "required": False,
                    "data_type": "string",
                },
                {
                    "name": "country_code3_gid",
                    "label": "Country Code3 GID",
                    "target_column": "COUNTRY_CODE3_GID",
                    "required": False,
                    "data_type": "string",
                },
                {
                    "name": "lat",
                    "label": "Latitude",
                    "target_column": "LAT",
                    "required": False,
                    "data_type": "number",
                },
                {
                    "name": "lon",
                    "label": "Longitude",
                    "target_column": "LON",
                    "required": False,
                    "data_type": "number",
                },
            ],
        },
        {
            "code": "LOCATION_ADDRESSES",
            "name": "Location Addresses",
            "target_table": "LOCATION_ADDRESS",
            "fields": [
                {
                    "name": "location_gid",
                    "label": "Location GID",
                    "target_column": "LOCATION_GID",
                    "required": True,
                    "data_type": "string",
                },
                {
                    "name": "line_sequence",
                    "label": "Line Sequence",
                    "target_column": "LINE_SEQUENCE",
                    "required": True,
                    "data_type": "number",
                },
                {
                    "name": "address_line",
                    "label": "Address Line",
                    "target_column": "ADDRESS_LINE",
                    "required": False,
                    "data_type": "string",
                },
            ],
        },
    ],
    "description": "Synthetic starter template for location master data and address lines.",
}

MASTER_DATA_TEMPLATE_SEEDS = [
    REGIONS_BASIC_TEMPLATE,
    ITEMS_PACKAGING_STANDARD_TEMPLATE,
    LOCATIONS_BASIC_TEMPLATE,
]

MASTER_DATA_TEMPLATE_DEFINITION_SCHEMA_VERSION = "master-data-template-definition/v2"

MASTER_DATA_RELATIONSHIP_RULES = {
    "REGIONS_BASIC": [
        {
            "parent_sheet_code": "REGIONS",
            "parent_field_name": "region_gid",
            "child_sheet_code": "REGION_DETAILS",
            "child_field_name": "region_gid",
        },
    ],
    "ITEMS_PACKAGING_STANDARD": [
        {
            "parent_sheet_code": "ITEMS",
            "parent_field_name": "item_gid",
            "child_sheet_code": "PACKAGING",
            "child_field_name": "item_gid",
        },
        {
            "parent_sheet_code": "PACKAGING",
            "parent_field_name": "packaged_item_gid",
            "child_sheet_code": "TI_HI",
            "child_field_name": "packaged_item_gid",
        },
    ],
    "LOCATIONS_BASIC": [
        {
            "parent_sheet_code": "LOCATIONS",
            "parent_field_name": "location_gid",
            "child_sheet_code": "LOCATION_ADDRESSES",
            "child_field_name": "location_gid",
        },
    ],
}


def seed_master_data_templates(db: Session) -> None:
    changed = False
    for seed in MASTER_DATA_TEMPLATE_SEEDS:
        exists = db.query(MasterDataTemplate).filter(MasterDataTemplate.code == seed["code"]).first()
        if exists:
            if not json.loads(exists.sheets_json):
                exists.sheets_json = json.dumps(seed["sheets"])
                changed = True
            continue
        db.add(
            MasterDataTemplate(
                code=seed["code"],
                name=seed["name"],
                version=seed["version"],
                status=seed["status"],
                catalog_macro_object_code=seed["catalog_macro_object_code"],
                data_category=seed["data_category"],
                target_tables_json=json.dumps(seed["target_tables"]),
                sheets_json=json.dumps(seed["sheets"]),
                description=seed["description"],
            )
        )
        changed = True
    if changed:
        db.commit()


def master_data_template_definition_from_seed(seed: dict[str, object]) -> dict[str, object]:
    target_tables = [
        {"table_name": table_name, "sequence": (index + 1) * 10, "required": True}
        for index, table_name in enumerate(seed["target_tables"])
    ]
    fields = []
    mappings = []
    sheets = []
    for sheet_index, sheet in enumerate(seed["sheets"]):
        field_keys = []
        for field in sheet["fields"]:
            field_key = field["name"]
            field_keys.append(field_key)
            fields.append(
                {
                    "field_key": field_key,
                    "label": field["label"],
                    "data_type": field.get("data_type", "string"),
                    "required": bool(field.get("required", False)),
                    "sheet_code": sheet["code"],
                }
            )
            mappings.append(
                {
                    "mapping_key": f"{sheet['code'].lower()}_{field_key}_to_{field['target_column'].lower()}",
                    "source_type": "USER_FIELD",
                    "sheet_code": sheet["code"],
                    "source_field_key": field_key,
                    "target_table": sheet["target_table"],
                    "target_column": field["target_column"],
                    "required": bool(field.get("required", False)),
                }
            )
        sheets.append(
            {
                "code": sheet["code"],
                "name": sheet["name"],
                "sequence": (sheet_index + 1) * 10,
                "field_keys": field_keys,
            }
        )

    return {
        "schema_version": MASTER_DATA_TEMPLATE_DEFINITION_SCHEMA_VERSION,
        "template": {
            "code": seed["code"],
            "name": seed["name"],
            "version": seed["version"],
            "status": seed["status"],
            "catalog_macro_object_code": seed["catalog_macro_object_code"],
            "data_category": seed["data_category"],
        },
        "target_tables": target_tables,
        "sheets": sheets,
        "fields": fields,
        "mappings": mappings,
        "relationship_rules": MASTER_DATA_RELATIONSHIP_RULES.get(seed["code"], []),
        "documentation_refs": [
            {
                "source_type": "DATA_DICTIONARY",
                "scope": seed["catalog_macro_object_code"],
                "note": "Validated against local OTM Data Dictionary.",
            }
        ],
    }


def master_data_template_definition(template: MasterDataTemplate) -> dict[str, object]:
    if template.definition_json and template.definition_json != "{}":
        return json.loads(template.definition_json)
    seed = {
        "code": template.code,
        "name": template.name,
        "version": template.version,
        "status": template.status,
        "catalog_macro_object_code": template.catalog_macro_object_code,
        "data_category": template.data_category,
        "target_tables": json.loads(template.target_tables_json),
        "sheets": json.loads(template.sheets_json),
    }
    return master_data_template_definition_from_seed(seed)


def normalize_master_data_template_definition(payload: dict[str, object], status: str = "DRAFT") -> dict[str, object]:
    code = str(payload["code"]).strip().upper()
    target_tables = []
    for index, table in enumerate(payload.get("target_tables", [])):
        table_name = str(table["table_name"]).strip().upper()
        target_tables.append(
            {
                "table_name": table_name,
                "sequence": int(table.get("sequence", (index + 1) * 10)),
                "required": bool(table.get("required", True)),
            }
        )
    sheets = []
    for index, sheet in enumerate(payload.get("sheets", [])):
        sheets.append(
            {
                "code": str(sheet["code"]).strip().upper(),
                "name": str(sheet["name"]).strip(),
                "sequence": int(sheet.get("sequence", (index + 1) * 10)),
                "field_keys": [str(field_key).strip() for field_key in sheet.get("field_keys", [])],
            }
        )
    fields = []
    for field in payload.get("fields", []):
        fields.append(
            {
                "field_key": str(field["field_key"]).strip(),
                "label": str(field["label"]).strip(),
                "data_type": str(field.get("data_type", "string")).strip().lower(),
                "required": bool(field.get("required", False)),
                "sheet_code": str(field["sheet_code"]).strip().upper(),
            }
        )
    mappings = []
    for mapping in payload.get("mappings", []):
        normalized_mapping = {
            "mapping_key": str(mapping["mapping_key"]).strip(),
            "source_type": str(mapping["source_type"]).strip().upper(),
            "target_table": str(mapping["target_table"]).strip().upper(),
            "target_column": str(mapping["target_column"]).strip().upper(),
            "required": bool(mapping.get("required", False)),
        }
        if mapping.get("source_field_key") is not None:
            normalized_mapping["source_field_key"] = str(mapping["source_field_key"]).strip()
        if mapping.get("sheet_code") is not None:
            normalized_mapping["sheet_code"] = str(mapping["sheet_code"]).strip().upper()
        if mapping.get("fixed_value") is not None:
            normalized_mapping["fixed_value"] = mapping["fixed_value"]
        if mapping.get("default_value") is not None:
            normalized_mapping["default_value"] = mapping["default_value"]
        mappings.append(normalized_mapping)

    return {
        "schema_version": MASTER_DATA_TEMPLATE_DEFINITION_SCHEMA_VERSION,
        "template": {
            "code": code,
            "name": str(payload["name"]).strip(),
            "version": int(payload.get("version", 1)),
            "status": status,
            "catalog_macro_object_code": str(payload["catalog_macro_object_code"]).strip().upper(),
            "data_category": str(payload.get("data_category", "MASTER_DATA")).strip().upper(),
        },
        "target_tables": target_tables,
        "sheets": sheets,
        "fields": fields,
        "mappings": mappings,
        "relationship_rules": payload.get("relationship_rules", []),
        "documentation_refs": payload.get("documentation_refs", []),
    }


def v2_definition_to_legacy_sheets(definition: dict[str, object]) -> list[dict[str, object]]:
    fields_by_key = {field["field_key"]: field for field in definition.get("fields", [])}
    first_mapping_by_field: dict[str, dict[str, object]] = {}
    for mapping in definition.get("mappings", []):
        if mapping.get("source_type") != "USER_FIELD":
            continue
        field_key = mapping.get("source_field_key")
        if field_key and field_key not in first_mapping_by_field:
            first_mapping_by_field[field_key] = mapping

    sheets = []
    for sheet in definition.get("sheets", []):
        legacy_fields = []
        target_table = None
        for field_key in sheet.get("field_keys", []):
            field = fields_by_key.get(field_key)
            mapping = first_mapping_by_field.get(field_key)
            if field is None or mapping is None:
                continue
            target_table = target_table or mapping["target_table"]
            legacy_fields.append(
                {
                    "name": field["field_key"],
                    "label": field["label"],
                    "target_column": mapping["target_column"],
                    "required": bool(field.get("required", False)),
                    "data_type": field.get("data_type", "string"),
                }
            )
        sheets.append(
            {
                "code": sheet["code"],
                "name": sheet["name"],
                "target_table": target_table or definition["target_tables"][0]["table_name"],
                "fields": legacy_fields,
            }
        )
    return sheets


def create_master_data_template_draft(
    db: Session,
    payload: dict[str, object],
) -> dict[str, object]:
    definition = normalize_master_data_template_definition(payload, status="DRAFT")
    template_meta = definition["template"]
    code = template_meta["code"]
    existing = db.query(MasterDataTemplate).filter(MasterDataTemplate.code == code).first()
    if existing is not None:
        raise ValueError("Master Data template code already exists.")

    target_tables = [table["table_name"] for table in definition["target_tables"]]
    legacy_sheets = v2_definition_to_legacy_sheets(definition)
    template = MasterDataTemplate(
        code=code,
        name=template_meta["name"],
        version=template_meta["version"],
        status="DRAFT",
        catalog_macro_object_code=template_meta["catalog_macro_object_code"],
        data_category=template_meta["data_category"],
        target_tables_json=json.dumps(target_tables),
        sheets_json=json.dumps(legacy_sheets),
        definition_json=json.dumps(definition, sort_keys=True),
        description=str(payload.get("description", "")),
    )
    db.add(template)
    db.commit()
    db.refresh(template)
    return serialize_master_data_template(template)


def update_master_data_template_draft(
    db: Session,
    template: MasterDataTemplate,
    payload: dict[str, object],
) -> dict[str, object]:
    if template.status != "DRAFT":
        raise ValueError("Only draft Master Data templates can be updated.")
    definition = normalize_master_data_template_definition(payload, status="DRAFT")
    template_meta = definition["template"]
    if template_meta["code"] != template.code:
        raise ValueError("Draft template code cannot be changed.")

    target_tables = [table["table_name"] for table in definition["target_tables"]]
    template.name = template_meta["name"]
    template.version = template_meta["version"]
    template.catalog_macro_object_code = template_meta["catalog_macro_object_code"]
    template.data_category = template_meta["data_category"]
    template.target_tables_json = json.dumps(target_tables)
    template.sheets_json = json.dumps(v2_definition_to_legacy_sheets(definition))
    template.definition_json = json.dumps(definition, sort_keys=True)
    template.description = str(payload.get("description", template.description))
    db.commit()
    db.refresh(template)
    return serialize_master_data_template(template)


def create_master_data_template_version(
    db: Session,
    source_template: MasterDataTemplate,
    new_code: str | None = None,
) -> dict[str, object]:
    code = (new_code or f"{source_template.code}_V{source_template.version + 1}").strip().upper()
    existing = db.query(MasterDataTemplate).filter(MasterDataTemplate.code == code).first()
    if existing is not None:
        raise ValueError("Master Data template version code already exists.")

    definition = master_data_template_definition(source_template)
    definition = json.loads(json.dumps(definition))
    definition["template"]["code"] = code
    definition["template"]["version"] = int(source_template.version) + 1
    definition["template"]["status"] = "DRAFT"

    target_tables = [table["table_name"] for table in definition["target_tables"]]
    template = MasterDataTemplate(
        code=code,
        name=source_template.name,
        version=int(source_template.version) + 1,
        status="DRAFT",
        catalog_macro_object_code=source_template.catalog_macro_object_code,
        data_category=source_template.data_category,
        target_tables_json=json.dumps(target_tables),
        sheets_json=json.dumps(v2_definition_to_legacy_sheets(definition)),
        definition_json=json.dumps(definition, sort_keys=True),
        description=source_template.description,
    )
    db.add(template)
    db.commit()
    db.refresh(template)
    return serialize_master_data_template(template)


def validate_master_data_template_definition(
    template: MasterDataTemplate,
    dictionary_root: Path,
) -> dict[str, object]:
    definition = master_data_template_definition(template)
    issues = []
    target_tables = {table["table_name"] for table in definition.get("target_tables", [])}
    field_keys = {field["field_key"] for field in definition.get("fields", [])}
    sheet_codes = {sheet["code"] for sheet in definition.get("sheets", [])}
    documentation_refs = definition.get("documentation_refs", [])
    allowed_documentation_sources = {"DATA_DICTIONARY", "ORACLE_OFFICIAL", "USER_CONFIRMED"}

    if not documentation_refs:
        issues.append(
            {
                "code": "MASTER_DATA_TEMPLATE_DOCUMENTATION_REF_REQUIRED",
                "severity": "ERROR",
                "message": "Dynamic Master Data templates must record Data Dictionary, Oracle official, or user-confirmed documentation references.",
            }
        )
    for reference in documentation_refs:
        source_type = str(reference.get("source_type", "")).strip().upper()
        if source_type not in allowed_documentation_sources:
            issues.append(
                {
                    "code": "MASTER_DATA_TEMPLATE_DOCUMENTATION_REF_INVALID",
                    "severity": "ERROR",
                    "source_type": source_type,
                    "message": "Documentation reference source_type must be DATA_DICTIONARY, ORACLE_OFFICIAL, or USER_CONFIRMED.",
                }
            )

    validated_tables = set()
    for table_name in target_tables:
        table_result = validate_table(dictionary_root, table_name, usage="cutover")
        if table_result["exists"] and table_result["severity"] != "ERROR":
            validated_tables.add(table_name)
        else:
            issues.append(
                {
                    "code": "MASTER_DATA_TEMPLATE_TARGET_TABLE_INVALID",
                    "severity": "ERROR",
                    "target_table": table_name,
                    "message": table_result["message"],
                }
            )

    seen_targets = set()
    for mapping in definition.get("mappings", []):
        target_table = mapping["target_table"]
        target_column = mapping["target_column"]
        target_key = (target_table, target_column)
        if target_key in seen_targets:
            issues.append(
                {
                    "code": "MASTER_DATA_TEMPLATE_MAPPING_TARGET_AMBIGUOUS",
                    "severity": "ERROR",
                    "target_table": target_table,
                    "target_column": target_column,
                    "message": "Target column has more than one unconditional mapping.",
                }
            )
        seen_targets.add(target_key)

        if target_table not in target_tables:
            issues.append(
                {
                    "code": "MASTER_DATA_TEMPLATE_MAPPING_TABLE_NOT_DECLARED",
                    "severity": "ERROR",
                    "target_table": target_table,
                    "message": "Mapping target table must be declared in target_tables.",
                }
            )
            continue
        column_result = validate_column(dictionary_root, target_table, target_column)
        if not column_result["exists"]:
            issues.append(
                {
                    "code": "MASTER_DATA_TEMPLATE_MAPPING_COLUMN_INVALID",
                    "severity": "ERROR",
                    "target_table": target_table,
                    "target_column": target_column,
                    "message": column_result["message"],
                }
            )
        source_type = mapping["source_type"]
        if source_type == "USER_FIELD" and mapping.get("source_field_key") not in field_keys:
            issues.append(
                {
                    "code": "MASTER_DATA_TEMPLATE_MAPPING_FIELD_INVALID",
                    "severity": "ERROR",
                    "mapping_key": mapping["mapping_key"],
                    "source_field_key": mapping.get("source_field_key"),
                    "message": "Mapping references an unknown template field.",
                }
            )
        if source_type in {"FIXED_VALUE", "DEFAULT_VALUE"} and (
            mapping.get("fixed_value") is None and mapping.get("default_value") is None
        ):
            issues.append(
                {
                    "code": "MASTER_DATA_TEMPLATE_MAPPING_VALUE_REQUIRED",
                    "severity": "ERROR",
                    "mapping_key": mapping["mapping_key"],
                    "message": "Fixed/default mappings must include a value.",
                }
            )

    for field in definition.get("fields", []):
        if field["sheet_code"] not in sheet_codes:
            issues.append(
                {
                    "code": "MASTER_DATA_TEMPLATE_FIELD_SHEET_INVALID",
                    "severity": "ERROR",
                    "field_key": field["field_key"],
                    "sheet_code": field["sheet_code"],
                    "message": "Field references an unknown sheet.",
                }
            )

    for sheet in definition.get("sheets", []):
        for field_key in sheet.get("field_keys", []):
            if field_key not in field_keys:
                issues.append(
                    {
                        "code": "MASTER_DATA_TEMPLATE_SHEET_FIELD_INVALID",
                        "severity": "ERROR",
                        "sheet_code": sheet["code"],
                        "field_key": field_key,
                        "message": "Sheet references an unknown field.",
                    }
                )

    return {
        "template_code": template.code,
        "valid": not issues,
        "severity": "INFO" if not issues else "ERROR",
        "issues": issues,
        "summary": {
            "target_table_count": len(target_tables),
            "validated_table_count": len(validated_tables),
            "sheet_count": len(definition.get("sheets", [])),
            "field_count": len(definition.get("fields", [])),
            "mapping_count": len(definition.get("mappings", [])),
        },
    }


def publish_master_data_template(
    db: Session,
    template: MasterDataTemplate,
    dictionary_root: Path,
) -> dict[str, object]:
    validation = validate_master_data_template_definition(template, dictionary_root)
    if not validation["valid"]:
        raise ValueError(json.dumps(validation, sort_keys=True))
    definition = master_data_template_definition(template)
    definition["template"]["status"] = "PUBLISHED"
    template.status = "PUBLISHED"
    template.definition_json = json.dumps(definition, sort_keys=True)
    db.commit()
    db.refresh(template)
    payload = serialize_master_data_template(template)
    payload["validation"] = validation
    return payload


def master_data_template_action(
    *,
    template_code: str,
    key: str,
    label: str,
    path_suffix: str,
    variant: str,
    icon_key: str,
    disabled: bool,
    disabled_reason: str | None,
    result_hint: str,
) -> dict[str, object]:
    return {
        "key": key,
        "label": label,
        "method": "POST",
        "href": f"/api/v1/modules/master-data/templates/{template_code}{path_suffix}",
        "variant": variant,
        "icon_key": icon_key,
        "requires_confirmation": False,
        "disabled": disabled,
        "disabled_reason": disabled_reason,
        "recommended": False,
        "permission": f"master_data.template.{key}",
        "result_hint": result_hint,
    }


def mark_recommended_template_action(
    actions: list[dict[str, object]],
    recommended_key: str,
) -> list[dict[str, object]]:
    for action in actions:
        action["recommended"] = action["key"] == recommended_key
    return actions


def build_master_data_template_available_actions(template: MasterDataTemplate) -> list[dict[str, object]]:
    is_published = template.status == "PUBLISHED"
    recommended_key = "build_workbook" if is_published else "validate_definition"
    actions = [
        master_data_template_action(
            template_code=template.code,
            key="validate_definition",
            label="Validate definition",
            path_suffix="/validate-definition",
            variant="secondary",
            icon_key="check-circle",
            disabled=False,
            disabled_reason=None,
            result_hint="refresh_object",
        ),
        master_data_template_action(
            template_code=template.code,
            key="publish_template",
            label="Publish template",
            path_suffix="/publish",
            variant="secondary",
            icon_key="upload-cloud",
            disabled=is_published,
            disabled_reason="TEMPLATE_ALREADY_PUBLISHED" if is_published else None,
            result_hint="refresh_object",
        ),
        master_data_template_action(
            template_code=template.code,
            key="build_workbook",
            label="Build workbook",
            path_suffix="/build-workbook",
            variant="secondary",
            icon_key="file-spreadsheet",
            disabled=not is_published,
            disabled_reason=None if is_published else "PUBLISHED_TEMPLATE_REQUIRED",
            result_hint="refresh_object",
        ),
        master_data_template_action(
            template_code=template.code,
            key="create_version",
            label="Create next version",
            path_suffix="/versions",
            variant="secondary",
            icon_key="copy",
            disabled=not is_published,
            disabled_reason=None if is_published else "PUBLISHED_TEMPLATE_REQUIRED",
            result_hint="refresh_object",
        ),
    ]
    return mark_recommended_template_action(actions, recommended_key)


def serialize_master_data_template(template: MasterDataTemplate) -> dict[str, object]:
    return {
        "id": template.id,
        "code": template.code,
        "name": template.name,
        "version": template.version,
        "status": template.status,
        "catalog_macro_object_code": template.catalog_macro_object_code,
        "data_category": template.data_category,
        "target_tables": json.loads(template.target_tables_json),
        "sheets": json.loads(template.sheets_json),
        "definition": master_data_template_definition(template),
        "available_actions": build_master_data_template_available_actions(template),
        "description": template.description,
        "created_at": template.created_at.isoformat() if template.created_at else None,
        "updated_at": template.updated_at.isoformat() if template.updated_at else None,
    }


def validate_master_data_template(template: MasterDataTemplate, dictionary_root: Path) -> dict[str, object]:
    sheets = json.loads(template.sheets_json)
    issues = []
    validated_tables = set()
    validated_column_count = 0
    field_count = 0
    for sheet in sheets:
        target_table = sheet["target_table"]
        table_result = validate_table(dictionary_root, target_table, usage="cutover")
        if table_result["exists"] and table_result["severity"] != "ERROR":
            validated_tables.add(target_table)
        else:
            issues.append(
                {
                    "code": "MASTER_DATA_TEMPLATE_TABLE_INVALID",
                    "severity": "ERROR",
                    "sheet_code": sheet["code"],
                    "target_table": target_table,
                    "message": table_result["message"],
                }
            )
        for field in sheet["fields"]:
            field_count += 1
            column_result = validate_column(dictionary_root, target_table, field["target_column"])
            if column_result["exists"]:
                validated_column_count += 1
            else:
                issues.append(
                    {
                        "code": "MASTER_DATA_TEMPLATE_COLUMN_INVALID",
                        "severity": "ERROR",
                        "sheet_code": sheet["code"],
                        "field_name": field["name"],
                        "target_table": target_table,
                        "target_column": field["target_column"],
                        "message": column_result["message"],
                    }
                )
    return {
        "template_code": template.code,
        "valid": not issues,
        "severity": "INFO" if not issues else "ERROR",
        "issues": issues,
        "summary": {
            "sheet_count": len(sheets),
            "field_count": field_count,
            "validated_table_count": len(validated_tables),
            "validated_column_count": validated_column_count,
        },
    }


def build_master_data_template_workbook(
    db: Session,
    template: MasterDataTemplate,
    artifact_root: Path,
) -> dict[str, object]:
    sheets = json.loads(template.sheets_json)
    workbook = Workbook()
    default_sheet = workbook.active
    field_count = 0

    for index, sheet in enumerate(sheets):
        worksheet = default_sheet if index == 0 else workbook.create_sheet()
        worksheet.title = sheet["code"][:31]
        fields = sheet["fields"]
        field_count += len(fields)
        worksheet.append([field["label"] for field in fields])
        worksheet.append([field["name"] for field in fields])
        worksheet.append([field["target_column"] for field in fields])

    output_dir = artifact_root / "master_data" / "templates" / template.code.lower()
    output_dir.mkdir(parents=True, exist_ok=True)
    file_name = f"{template.code.lower()}_v{template.version}.xlsx"
    output_path = output_dir / file_name
    workbook.save(output_path)

    digest, size = file_sha256(str(output_path))
    artifact = Artifact(
        source_module="master_data",
        artifact_type="master_data_template_workbook",
        file_path=str(output_path),
        file_name=file_name,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        sha256=digest,
        size_bytes=size,
        sensitivity_level="client_safe",
    )
    db.add(artifact)
    db.commit()
    db.refresh(artifact)

    return {
        "template_code": template.code,
        "artifact_id": artifact.id,
        "file_name": artifact.file_name,
        "content_type": artifact.content_type,
        "sheet_count": len(sheets),
        "field_count": field_count,
    }


def parse_master_data_template_workbook(
    db: Session,
    template: MasterDataTemplate,
    file_obj: BinaryIO,
    file_name: str,
    content_type: str,
) -> dict[str, object]:
    sheets = json.loads(template.sheets_json)
    workbook = load_workbook(file_obj, data_only=True)
    sheet_summaries = []
    parsed_rows: dict[str, list[dict[str, object]]] = {}
    issues = []
    row_count = 0

    for sheet in sheets:
        if sheet["code"] not in workbook.sheetnames:
            issues.append(
                {
                    "code": "MASTER_DATA_WORKBOOK_SHEET_MISSING",
                    "severity": "ERROR",
                    "sheet_code": sheet["code"],
                    "message": "Uploaded workbook is missing a template sheet.",
                }
            )
            continue
        worksheet = workbook[sheet["code"]]
        fields = sheet["fields"]
        expected_headers = [field["label"] for field in fields]
        actual_headers = [cell.value for cell in worksheet[1]][: len(expected_headers)]
        if actual_headers != expected_headers:
            issues.append(
                {
                    "code": "MASTER_DATA_WORKBOOK_HEADERS_INVALID",
                    "severity": "ERROR",
                    "sheet_code": sheet["code"],
                    "message": "Uploaded workbook headers do not match the template.",
                    "expected_headers": expected_headers,
                    "actual_headers": actual_headers,
                }
            )
            continue

        sheet_rows = []
        expected_field_names = [field["name"] for field in fields]
        expected_target_columns = [field["target_column"] for field in fields]
        field_name_row = [cell.value for cell in worksheet[2]][: len(expected_field_names)]
        target_column_row = [cell.value for cell in worksheet[3]][: len(expected_target_columns)]
        data_start_row = 4 if (
            field_name_row == expected_field_names
            and target_column_row == expected_target_columns
        ) else 2
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=data_start_row, values_only=True),
            start=data_start_row,
        ):
            values = list(row[: len(fields)])
            if all(value is None for value in values):
                continue
            parsed_row = {field["name"]: values[index] for index, field in enumerate(fields)}
            parsed_row["_row_number"] = row_number
            sheet_rows.append(parsed_row)

        parsed_rows[sheet["code"]] = sheet_rows
        sheet_row_count = len(sheet_rows)
        row_count += sheet_row_count
        sheet_summaries.append(
            {
                "sheet_code": sheet["code"],
                "target_table": sheet["target_table"],
                "row_count": sheet_row_count,
            }
        )

    batch = MasterDataBatch(
        template_id=template.id,
        template_code=template.code,
        status="PARSED",
        file_name=file_name,
        content_type=content_type,
        sheet_summaries_json=json.dumps(sheet_summaries),
        parsed_rows_json=json.dumps(parsed_rows, sort_keys=True),
        issues_json=json.dumps(issues, sort_keys=True),
        row_count=0 if issues else row_count,
        issue_count=len(issues),
    )
    if issues:
        batch.status = "PARSE_FAILED"
    db.add(batch)
    db.commit()
    db.refresh(batch)

    return {
        "batch_id": batch.id,
        "template_code": batch.template_code,
        "status": batch.status,
        "file_name": batch.file_name,
        "sheet_count": len(sheet_summaries),
        "row_count": batch.row_count,
        "issue_count": batch.issue_count,
        "sheet_summaries": sheet_summaries,
        "issues": issues,
    }


def validate_master_data_batch_relationships(
    db: Session,
    batch: MasterDataBatch,
) -> dict[str, object]:
    if batch.status in {"RELATIONSHIP_VALIDATED", "RELATIONSHIP_FAILED"}:
        issues = json.loads(batch.issues_json or "[]")
        return {
            "batch_id": batch.id,
            "status": batch.status,
            "issue_count": batch.issue_count,
            "issues": issues,
        }
    if batch.status != "PARSED":
        raise ValueError("Only parsed Master Data batches can be relationship-validated.")

    parsed_rows = json.loads(batch.parsed_rows_json)
    template = (
        db.query(MasterDataTemplate)
        .filter(MasterDataTemplate.code == batch.template_code)
        .first()
    )
    relationship_rules = MASTER_DATA_RELATIONSHIP_RULES.get(batch.template_code, [])
    if template is not None:
        relationship_rules = master_data_template_definition(template).get("relationship_rules", relationship_rules)
    issues = []
    for rule in relationship_rules:
        parent_field_name = rule.get("parent_field_name") or rule.get("parent_field_key")
        child_field_name = rule.get("child_field_name") or rule.get("child_field_key")
        parent_values = {
            row.get(parent_field_name)
            for row in parsed_rows.get(rule["parent_sheet_code"], [])
            if row.get(parent_field_name) not in {None, ""}
        }
        for row_index, child_row in enumerate(parsed_rows.get(rule["child_sheet_code"], []), start=2):
            child_value = child_row.get(child_field_name)
            if child_value in {None, ""} or child_value in parent_values:
                continue
            issue = {
                    "code": "MASTER_DATA_RELATIONSHIP_ORPHAN",
                    "severity": "ERROR",
                    "sheet_code": rule["child_sheet_code"],
                    "field_name": child_field_name,
                    "row_number": child_row.get("_row_number", row_index),
                    "message": "Child row references a parent key that does not exist in the same batch.",
                    "parent_sheet_code": rule["parent_sheet_code"],
                    "parent_field_name": parent_field_name,
                    "missing_value": child_value,
            }
            if "child_field_key" in rule:
                issue["child_field_name"] = child_field_name
            issues.append(issue)

    batch.status = "RELATIONSHIP_FAILED" if issues else "RELATIONSHIP_VALIDATED"
    batch.issues_json = json.dumps(issues, sort_keys=True)
    batch.issue_count = len(issues)
    db.commit()

    return {
        "batch_id": batch.id,
        "status": batch.status,
        "issue_count": batch.issue_count,
        "issues": issues,
    }


def map_master_data_batch_to_canonical_records(
    db: Session,
    template: MasterDataTemplate,
    batch: MasterDataBatch,
) -> dict[str, object]:
    if MASTER_DATA_RELATIONSHIP_RULES.get(batch.template_code):
        if batch.status != "RELATIONSHIP_VALIDATED":
            raise ValueError(
                "Master Data batch must be relationship-validated before mapping."
            )
    elif batch.status != "PARSED":
        raise ValueError("Only parsed Master Data batches can be mapped.")

    sheets = json.loads(template.sheets_json)
    definition = master_data_template_definition(template)
    definition_mappings = definition.get("mappings", [])
    definition_fields_by_sheet: dict[str, set[str]] = {}
    for field in definition.get("fields", []):
        definition_fields_by_sheet.setdefault(field["sheet_code"], set()).add(field["field_key"])
    parsed_rows = json.loads(batch.parsed_rows_json)
    db.query(MasterDataCanonicalRecord).filter(
        MasterDataCanonicalRecord.batch_id == batch.id
    ).delete()

    response_records = []
    for sheet in sheets:
        rows = parsed_rows.get(sheet["code"], [])
        for index, parsed_row in enumerate(rows, start=1):
            field_keys_for_sheet = definition_fields_by_sheet.get(sheet["code"], set())
            payloads_by_table: dict[str, dict[str, object]] = {}
            for mapping in definition_mappings:
                mapping_sheet_code = mapping.get("sheet_code")
                if mapping_sheet_code and mapping_sheet_code != sheet["code"]:
                    continue
                source_type = mapping["source_type"]
                source_field_key = mapping.get("source_field_key")
                if source_type == "USER_FIELD":
                    if source_field_key not in field_keys_for_sheet:
                        continue
                    value = parsed_row.get(source_field_key)
                elif source_type == "FIXED_VALUE":
                    value = mapping.get("fixed_value")
                elif source_type == "DEFAULT_VALUE":
                    if source_field_key and source_field_key not in field_keys_for_sheet:
                        continue
                    source_value = parsed_row.get(source_field_key) if source_field_key else None
                    value = source_value if source_value not in {None, ""} else mapping.get("default_value")
                else:
                    continue
                payloads_by_table.setdefault(mapping["target_table"], {})[mapping["target_column"]] = value

            for target_table, payload in payloads_by_table.items():
                canonical_record = MasterDataCanonicalRecord(
                    batch_id=batch.id,
                    template_code=batch.template_code,
                    sheet_code=sheet["code"],
                    target_table=target_table,
                    record_index=index,
                    payload_json=json.dumps(payload, sort_keys=True),
                )
                db.add(canonical_record)
                response_records.append(
                    {
                        "sheet_code": sheet["code"],
                        "target_table": target_table,
                        "record_index": index,
                        "payload": payload,
                    }
                )

    batch.status = "MAPPED"
    db.commit()

    return {
        "batch_id": batch.id,
        "status": batch.status,
        "canonical_record_count": len(response_records),
        "records": response_records,
    }


def build_master_data_output_records(
    db: Session,
    batch: MasterDataBatch,
) -> dict[str, object]:
    if batch.status != "MAPPED":
        raise ValueError("Only mapped Master Data batches can build output records.")

    db.query(MasterDataOutputRecord).filter(MasterDataOutputRecord.batch_id == batch.id).delete()
    canonical_records = (
        db.query(MasterDataCanonicalRecord)
        .filter(MasterDataCanonicalRecord.batch_id == batch.id)
        .order_by(MasterDataCanonicalRecord.created_at, MasterDataCanonicalRecord.record_index)
        .all()
    )

    response_records = []
    for canonical_record in canonical_records:
        payload = json.loads(canonical_record.payload_json)
        output_record = MasterDataOutputRecord(
            batch_id=batch.id,
            template_code=batch.template_code,
            target_table=canonical_record.target_table,
            record_index=canonical_record.record_index,
            payload_json=json.dumps(payload, sort_keys=True),
        )
        db.add(output_record)
        response_records.append(
            {
                "target_table": canonical_record.target_table,
                "record_index": canonical_record.record_index,
                "payload": payload,
            }
        )

    batch.status = "OUTPUT_BUILT"
    db.commit()

    return {
        "batch_id": batch.id,
        "status": batch.status,
        "output_record_count": len(response_records),
        "records": response_records,
    }


def build_master_data_csv_files(
    db: Session,
    batch: MasterDataBatch,
    dictionary_root: Path,
) -> dict[str, object]:
    if batch.status == "CSV_BUILT":
        csv_files = (
            db.query(MasterDataCsvFile)
            .filter(MasterDataCsvFile.batch_id == batch.id)
            .order_by(MasterDataCsvFile.file_name)
            .all()
        )
        if csv_files:
            return {
                "batch_id": batch.id,
                "status": batch.status,
                "csv_file_count": len(csv_files),
                "files": [
                    {
                        "table_name": csv_file.table_name,
                        "file_name": csv_file.file_name,
                        "row_count": csv_file.row_count,
                        "content": csv_file.content,
                    }
                    for csv_file in csv_files
                ],
            }
    if batch.status != "OUTPUT_BUILT":
        raise ValueError("Only output-built Master Data batches can build CSV files.")

    db.query(MasterDataCsvFile).filter(MasterDataCsvFile.batch_id == batch.id).delete()
    output_records = (
        db.query(MasterDataOutputRecord)
        .filter(MasterDataOutputRecord.batch_id == batch.id)
        .order_by(MasterDataOutputRecord.created_at, MasterDataOutputRecord.record_index)
        .all()
    )
    records_by_table: dict[str, list[dict[str, object]]] = {}
    for output_record in output_records:
        records_by_table.setdefault(output_record.target_table, []).append(
            json.loads(output_record.payload_json)
        )

    template = (
        db.query(MasterDataTemplate)
        .filter(MasterDataTemplate.code == batch.template_code)
        .first()
    )
    template_columns_by_table: dict[str, list[str]] = {}
    if template is not None:
        definition = master_data_template_definition(template)
        for mapping in definition.get("mappings", []):
            target_table = mapping["target_table"]
            table_columns = template_columns_by_table.setdefault(target_table, [])
            target_column = mapping["target_column"]
            if target_column not in table_columns:
                table_columns.append(target_column)
        if not template_columns_by_table:
            for sheet in json.loads(template.sheets_json):
                target_table = sheet["target_table"]
                table_columns = template_columns_by_table.setdefault(target_table, [])
                for field in sheet["fields"]:
                    target_column = field["target_column"]
                    if target_column not in table_columns:
                        table_columns.append(target_column)

    response_files = []
    for index, table_name in enumerate(records_by_table, start=1):
        rows = records_by_table[table_name]
        row_columns = {column for row in rows for column in row}
        template_columns = template_columns_by_table.get(table_name, [])
        columns = [column for column in template_columns if column in row_columns]
        columns.extend(sorted(row_columns - set(columns)))
        content = build_otm_csv_preview(dictionary_root, table_name, columns, rows)
        file_name = f"{index:03d}_{table_name}.csv"
        csv_file = MasterDataCsvFile(
            batch_id=batch.id,
            template_code=batch.template_code,
            table_name=table_name,
            file_name=file_name,
            row_count=len(rows),
            content=content,
        )
        db.add(csv_file)
        response_files.append(
            {
                "table_name": table_name,
                "file_name": file_name,
                "row_count": len(rows),
                "content": content,
            }
        )

    batch.status = "CSV_BUILT"
    db.commit()

    return {
        "batch_id": batch.id,
        "status": batch.status,
        "csv_file_count": len(response_files),
        "files": response_files,
    }


def export_master_data_csv_package(
    db: Session,
    batch: MasterDataBatch,
    artifact_root: Path,
    generated_by: str,
) -> dict[str, object]:
    csv_files = (
        db.query(MasterDataCsvFile)
        .filter(MasterDataCsvFile.batch_id == batch.id)
        .order_by(MasterDataCsvFile.file_name)
        .all()
    )
    if batch.status == "EXPORTED":
        existing_evidence = (
            db.query(Evidence)
            .filter(Evidence.source_module == "master_data")
            .filter(Evidence.evidence_type == "master_data_csv_export")
            .filter(Evidence.client_safe.is_(True))
            .order_by(Evidence.created_at.desc())
            .all()
        )
        for evidence in existing_evidence:
            summary = json.loads(evidence.summary_json or "{}")
            if summary.get("source_entity_id") != batch.id:
                continue
            artifact = db.query(Artifact).filter(Artifact.id == evidence.artifact_id).first()
            manifest = db.query(Manifest).filter(Manifest.id == evidence.manifest_id).first()
            if artifact is None or manifest is None:
                continue
            return {
                "batch_id": batch.id,
                "status": batch.status,
                "artifact_id": artifact.id,
                "manifest_id": manifest.id,
                "evidence_id": evidence.id,
                "file_name": artifact.file_name,
                "sha256": artifact.sha256,
                "size_bytes": artifact.size_bytes,
                "tables": [csv_file.table_name for csv_file in csv_files],
            }
    if batch.status != "CSV_BUILT":
        raise ValueError("Only CSV-built Master Data batches can be exported.")

    if not csv_files:
        raise ValueError("Master Data batch has no CSV files to export.")

    timestamp = datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")
    export_dir = artifact_root / "master_data" / batch.id / "csv_exports" / timestamp
    export_dir.mkdir(parents=True, exist_ok=True)
    zip_path = export_dir / f"master_data_batch_{batch.id}.zip"

    manifest_files = [
        {
            "table_name": csv_file.table_name,
            "file_name": csv_file.file_name,
            "row_count": csv_file.row_count,
            "size_bytes": len(csv_file.content.encode("utf-8")),
        }
        for csv_file in csv_files
    ]
    manifest_payload = {
        "schema_version": "master-data-csv-export-manifest/v1",
        "manifest_type": "master_data_csv_export",
        "source_module": "master_data",
        "source_entity_type": "master_data_batch",
        "source_entity_id": batch.id,
        "batch": {
            "id": batch.id,
            "template_code": batch.template_code,
            "status": "EXPORTED",
            "row_count": batch.row_count,
            "issue_count": batch.issue_count,
        },
        "files": manifest_files,
        "generated_at": datetime.now(UTC).replace(microsecond=0).isoformat(),
        "generated_by": generated_by,
    }
    manifest_text = json.dumps(manifest_payload, indent=2, sort_keys=True)

    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("manifest.json", manifest_text)
        for csv_file in csv_files:
            archive.writestr(f"csv/{csv_file.file_name}", csv_file.content)

    zip_hash, zip_size = file_sha256(str(zip_path))
    artifact = Artifact(
        source_module="master_data",
        artifact_type="master_data_csv_zip",
        file_path=str(zip_path),
        file_name=zip_path.name,
        content_type="application/zip",
        sha256=zip_hash,
        size_bytes=zip_size,
        sensitivity_level="client_safe",
    )
    db.add(artifact)
    db.flush()

    manifest = Manifest(
        source_module="master_data",
        status="CREATED",
        manifest_json=manifest_text,
    )
    db.add(manifest)
    db.flush()

    evidence_summary = {
        "source_entity_type": "master_data_batch",
        "source_entity_id": batch.id,
        "template_code": batch.template_code,
        "table_count": len(csv_files),
        "row_count": sum(csv_file.row_count for csv_file in csv_files),
        "artifact_type": "master_data_csv_zip",
    }
    evidence = Evidence(
        source_module="master_data",
        evidence_type="master_data_csv_export",
        summary_json=json.dumps(evidence_summary, sort_keys=True),
        artifact_id=artifact.id,
        manifest_id=manifest.id,
        client_safe=True,
        sensitivity_level="client_safe",
    )
    db.add(evidence)
    db.flush()

    audit = AuditLog(
        actor_user_id=generated_by,
        action="master_data.batch.export_csv",
        target_type="master_data_batch",
        target_id=batch.id,
        metadata_json=json.dumps(
            {
                "artifact_id": artifact.id,
                "manifest_id": manifest.id,
                "evidence_id": evidence.id,
                "table_count": len(csv_files),
            },
            sort_keys=True,
        ),
    )
    db.add(audit)

    batch.status = "EXPORTED"
    db.commit()

    return {
        "batch_id": batch.id,
        "status": batch.status,
        "artifact_id": artifact.id,
        "manifest_id": manifest.id,
        "evidence_id": evidence.id,
        "file_name": artifact.file_name,
        "sha256": artifact.sha256,
        "size_bytes": artifact.size_bytes,
        "tables": [csv_file.table_name for csv_file in csv_files],
    }
