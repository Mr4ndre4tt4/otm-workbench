import json
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook
from xml.etree import ElementTree


FIXTURE_ROOT = Path(__file__).parent / "fixtures" / "synthetic"


def test_synthetic_fixture_manifest_references_existing_files():
    manifest = json.loads((FIXTURE_ROOT / "manifest.json").read_text(encoding="utf-8"))

    assert manifest["scope"] == {
        "project": "SYNTHETIC_PROJECT",
        "environment": "UAT",
        "domain_name": "OTM1",
        "visibility": "PRIVATE",
    }
    for item in manifest["files"]:
        assert (FIXTURE_ROOT / item["path"]).exists(), item["path"]


def test_synthetic_otm_csv_preserves_csvutil_shape():
    lines = (FIXTURE_ROOT / "rates" / "rate_geo_cost.csv").read_text(encoding="utf-8").splitlines()

    assert lines[0] == "RATE_GEO_COST"
    assert lines[1].split(",") == [
        "RATE_GEO_COST_GROUP_GID",
        "RATE_GEO_COST_SEQ",
        "DESCRIPTION",
        "CHARGE_AMOUNT",
        "CHARGE_CURRENCY_GID",
        "CHARGE_TYPE",
        "COST_TYPE",
        "EFFECTIVE_DATE",
        "EXPIRATION_DATE",
        "IS_FILED_AS_TARIFF",
        "DOMAIN_NAME",
        "INSERT_USER",
        "INSERT_DATE",
    ]
    assert lines[2].startswith("exec alter session set nls_date_format")
    values = dict(zip(lines[1].split(","), lines[3].split(","), strict=True))
    assert values["DOMAIN_NAME"] == "OTM1"
    assert values["INSERT_DATE"] == "2026-01-01 00:00:00"


def test_synthetic_otm_csv_columns_exist_in_local_data_dictionary():
    data_dict_path = (
        Path(__file__).parents[1]
        / "OTM_RESOURCES"
        / "DATA_DICT26B"
        / "data_dictionary"
        / "json"
        / "data_dict"
        / "RATE_GEO_COST.json"
    )
    data_dict = json.loads(data_dict_path.read_text(encoding="utf-8"))
    data_dict_columns = {column["name"] for column in data_dict["columns"]}
    csv_columns = (FIXTURE_ROOT / "rates" / "rate_geo_cost.csv").read_text(encoding="utf-8").splitlines()[1].split(",")

    assert set(csv_columns).issubset(data_dict_columns)


def test_synthetic_payload_fixtures_are_parseable():
    ElementTree.parse(FIXTURE_ROOT / "integration" / "planned_shipment.xml")
    ElementTree.parse(FIXTURE_ROOT / "integration" / "shipment_to_delivery.xslt")

    delivery = json.loads((FIXTURE_ROOT / "integration" / "delivery_payload.json").read_text(encoding="utf-8"))
    order_release = json.loads(
        (FIXTURE_ROOT / "order-release" / "order_release_batch.json").read_text(encoding="utf-8")
    )

    assert delivery["domainName"] == "OTM1"
    assert order_release["rows"][0]["orderReleaseGid"] == "OTM1.OR_SYN_001"


def test_synthetic_binary_fixtures_are_valid_containers():
    workbook = load_workbook(FIXTURE_ROOT / "master-data" / "location_upload.xlsx")
    assert workbook.sheetnames == ["LOCATIONS", "LOCATION_ADDRESSES"]

    with ZipFile(FIXTURE_ROOT / "assets" / "client_safe_asset.docx") as docx:
        assert "[Content_Types].xml" in docx.namelist()
        assert "word/document.xml" in docx.namelist()

    pdf_bytes = (FIXTURE_ROOT / "load-plan" / "cutover_handoff.pdf").read_bytes()
    assert pdf_bytes.startswith(b"%PDF")
    assert b"%%EOF" in pdf_bytes[-32:]

    with ZipFile(FIXTURE_ROOT / "load-plan" / "csvutil_package.zip") as package:
        assert sorted(package.namelist()) == ["csv/RATE_GEO_COST.csv", "manifest.json"]
