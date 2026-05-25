import json
import zipfile
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl import load_workbook
from sqlalchemy import text

from otm_workbench.models import Artifact, AuditLog, Evidence, Manifest


def action_by_key(actions, key):
    return next(action for action in actions if action["key"] == key)


def test_master_data_health(client, admin_header):
    response = client.get("/api/v1/modules/master-data/health", headers=admin_header)

    assert response.status_code == 200
    assert response.json() == {
        "status": "ok",
        "module": "master_data",
        "catalog_macro_object_code": "REGION",
    }


def test_master_data_templates_seed_regions_basic(client, admin_header):
    response = client.get("/api/v1/modules/master-data/templates", headers=admin_header)

    assert response.status_code == 200
    payload = response.json()
    assert payload["total"] == 3
    templates_by_code = {template["code"]: template for template in payload["items"]}
    template = templates_by_code["REGIONS_BASIC"]
    assert template["code"] == "REGIONS_BASIC"
    assert template["name"] == "Regions Basic"
    assert template["version"] == 1
    assert template["status"] == "PUBLISHED"
    assert template["catalog_macro_object_code"] == "REGION"
    assert template["target_tables"] == ["REGION", "REGION_DETAIL"]
    assert template["data_category"] == "MASTER_DATA"
    items_template = templates_by_code["ITEMS_PACKAGING_STANDARD"]
    assert items_template["name"] == "Items & Packaging Standard"
    assert items_template["catalog_macro_object_code"] == "ITEM"
    assert items_template["target_tables"] == ["ITEM", "SHIP_UNIT_SPEC", "PACKAGED_ITEM", "TI_HI"]
    locations_template = templates_by_code["LOCATIONS_BASIC"]
    assert locations_template["name"] == "Locations Basic"
    assert locations_template["catalog_macro_object_code"] == "LOCATION"
    assert locations_template["target_tables"] == ["LOCATION", "LOCATION_ADDRESS"]


def test_master_data_template_detail_exposes_sheets_and_fields(client, admin_header):
    response = client.get("/api/v1/modules/master-data/templates/REGIONS_BASIC", headers=admin_header)

    assert response.status_code == 200
    payload = response.json()
    assert payload["code"] == "REGIONS_BASIC"
    assert payload["sheets"] == [
        {
            "code": "REGIONS",
            "name": "Regions",
            "target_table": "REGION",
            "fields": [
                {
                    "name": "region_gid",
                    "label": "Region GID",
                    "target_column": "REGION_GID",
                    "required": True,
                    "data_type": "string",
                },
                {
                    "name": "region_xid",
                    "label": "Region XID",
                    "target_column": "REGION_XID",
                    "required": True,
                    "data_type": "string",
                },
                {
                    "name": "region_name",
                    "label": "Region Name",
                    "target_column": "REGION_NAME",
                    "required": False,
                    "data_type": "string",
                },
            ],
        },
        {
            "code": "REGION_DETAILS",
            "name": "Region Details",
            "target_table": "REGION_DETAIL",
            "fields": [
                {
                    "name": "region_gid",
                    "label": "Region GID",
                    "target_column": "REGION_GID",
                    "required": True,
                    "data_type": "string",
                },
                {
                    "name": "location_gid",
                    "label": "Location GID",
                    "target_column": "LOCATION_GID",
                    "required": True,
                    "data_type": "string",
                },
            ],
        },
    ]


def test_master_data_template_detail_exposes_backend_owned_available_actions(client, admin_header):
    draft_response = client.post(
        "/api/v1/modules/master-data/templates/drafts",
        json=dynamic_locations_template_payload("LOCATIONS_TEMPLATE_ACTIONS"),
        headers=admin_header,
    )
    assert draft_response.status_code == 200
    draft_actions = draft_response.json()["available_actions"]
    assert action_by_key(draft_actions, "validate_definition")["disabled"] is False
    assert action_by_key(draft_actions, "validate_definition")["recommended"] is True
    assert action_by_key(draft_actions, "publish_template")["disabled"] is False
    assert action_by_key(draft_actions, "build_workbook")["disabled"] is True
    assert action_by_key(draft_actions, "build_workbook")["disabled_reason"] == "PUBLISHED_TEMPLATE_REQUIRED"
    assert action_by_key(draft_actions, "create_version")["disabled_reason"] == "PUBLISHED_TEMPLATE_REQUIRED"

    publish_response = client.post(
        "/api/v1/modules/master-data/templates/LOCATIONS_TEMPLATE_ACTIONS/publish",
        headers=admin_header,
    )
    assert publish_response.status_code == 200
    published_actions = publish_response.json()["available_actions"]
    assert action_by_key(published_actions, "publish_template")["disabled"] is True
    assert action_by_key(published_actions, "publish_template")["disabled_reason"] == "TEMPLATE_ALREADY_PUBLISHED"
    assert action_by_key(published_actions, "build_workbook")["disabled"] is False
    assert action_by_key(published_actions, "build_workbook")["recommended"] is True
    assert action_by_key(published_actions, "create_version")["disabled"] is False


def test_master_data_template_validation_uses_catalog_dictionary(client, admin_header):
    response = client.post("/api/v1/modules/master-data/templates/REGIONS_BASIC/validate", headers=admin_header)

    assert response.status_code == 200
    payload = response.json()
    assert payload["template_code"] == "REGIONS_BASIC"
    assert payload["valid"] is True
    assert payload["severity"] == "INFO"
    assert payload["issues"] == []
    assert payload["summary"] == {
        "sheet_count": 2,
        "field_count": 5,
        "validated_table_count": 2,
        "validated_column_count": 5,
    }


def test_master_data_scenario_packs_expose_backend_owned_draft_payloads(client, admin_header):
    response = client.get("/api/v1/modules/master-data/scenario-packs", headers=admin_header)

    assert response.status_code == 200
    payload = response.json()
    assert payload["total"] == 2
    packs = {pack["code"]: pack for pack in payload["items"]}

    location_pack = packs["LOCATION_OPERATIONAL"]
    assert location_pack["catalog_macro_object_code"] == "LOCATION"
    assert location_pack["target_tables"] == [
        "LOCATION",
        "LOCATION_ADDRESS",
        "LOCATION_CAPACITY",
        "LOCATION_ACTIVITY_TIME_DEF",
        "LOCATION_LOAD_UNLOAD_POINT",
        "EQUIPMENT_GROUP_PROFILE",
        "EQUIPMENT_GROUP_PROFILE_D",
    ]
    assert location_pack["summary"] == {
        "sheet_count": 7,
        "field_count": 34,
        "mapping_count": 33,
        "relationship_rule_count": 6,
    }
    assert location_pack["draft_payload"]["code"] == "LOCATIONS_OPERATIONAL_QA"
    assert location_pack["draft_payload"]["relationship_rules"][0]["rule_key"] == "location_address_parent"
    assert location_pack["draft_payload"]["documentation_refs"][1]["source_type"] == "ORACLE_OFFICIAL"

    item_pack = packs["ITEM_PACKAGING_OPERATIONAL"]
    assert item_pack["catalog_macro_object_code"] == "ITEM"
    assert item_pack["target_tables"] == ["ITEM", "SHIP_UNIT_SPEC", "PACKAGED_ITEM", "TI_HI"]
    assert item_pack["summary"]["relationship_rule_count"] == 5
    assert any(
        mapping["target_column"] == "TRANSPORT_HANDLING_UNIT_GID"
        for mapping in item_pack["draft_payload"]["mappings"]
        if mapping["target_table"] == "TI_HI"
    )


def dynamic_locations_template_payload(code: str = "LOCATIONS_DYNAMIC") -> dict[str, object]:
    return {
        "code": code,
        "name": "Locations Dynamic",
        "catalog_macro_object_code": "LOCATION",
        "data_category": "MASTER_DATA",
        "target_tables": [
            {"table_name": "LOCATION", "sequence": 10, "required": True},
        ],
        "sheets": [
            {
                "code": "LOCATIONS",
                "name": "Locations",
                "sequence": 10,
                "field_keys": ["location_gid", "location_name"],
            },
        ],
        "fields": [
            {
                "field_key": "location_gid",
                "label": "Location ID",
                "data_type": "string",
                "required": True,
                "sheet_code": "LOCATIONS",
            },
            {
                "field_key": "location_name",
                "label": "Location Name",
                "data_type": "string",
                "required": False,
                "sheet_code": "LOCATIONS",
            },
        ],
        "mappings": [
            {
                "mapping_key": "location_gid_to_gid",
                "source_type": "USER_FIELD",
                "source_field_key": "location_gid",
                "target_table": "LOCATION",
                "target_column": "LOCATION_GID",
                "required": True,
            },
            {
                "mapping_key": "location_gid_to_xid",
                "source_type": "USER_FIELD",
                "source_field_key": "location_gid",
                "target_table": "LOCATION",
                "target_column": "LOCATION_XID",
                "required": True,
            },
            {
                "mapping_key": "location_name_to_name",
                "source_type": "USER_FIELD",
                "source_field_key": "location_name",
                "target_table": "LOCATION",
                "target_column": "LOCATION_NAME",
                "required": False,
            },
            {
                "mapping_key": "country_default",
                "source_type": "FIXED_VALUE",
                "fixed_value": "USA",
                "target_table": "LOCATION",
                "target_column": "COUNTRY_CODE3_GID",
                "required": False,
            },
        ],
        "relationship_rules": [],
        "documentation_refs": [
            {
                "source_type": "DATA_DICTIONARY",
                "scope": "LOCATION",
                "note": "Validated against local OTM Data Dictionary.",
            },
        ],
    }


def dynamic_locations_with_addresses_payload(code: str = "LOCATIONS_DYNAMIC_REL") -> dict[str, object]:
    payload = dynamic_locations_template_payload(code)
    payload["target_tables"].append(
        {"table_name": "LOCATION_ADDRESS", "sequence": 20, "required": False}
    )
    payload["sheets"].append(
        {
            "code": "LOCATION_ADDRESSES",
            "name": "Location Addresses",
            "sequence": 20,
            "field_keys": ["address_location_gid", "address_line"],
        }
    )
    payload["fields"].extend(
        [
            {
                "field_key": "address_location_gid",
                "label": "Address Location ID",
                "data_type": "string",
                "required": True,
                "sheet_code": "LOCATION_ADDRESSES",
            },
            {
                "field_key": "address_line",
                "label": "Address Line",
                "data_type": "string",
                "required": False,
                "sheet_code": "LOCATION_ADDRESSES",
            },
        ]
    )
    payload["mappings"].extend(
        [
            {
                "mapping_key": "address_location_gid_to_location_address",
                "source_type": "USER_FIELD",
                "source_field_key": "address_location_gid",
                "target_table": "LOCATION_ADDRESS",
                "target_column": "LOCATION_GID",
                "required": True,
            },
            {
                "mapping_key": "address_line_to_location_address",
                "source_type": "USER_FIELD",
                "source_field_key": "address_line",
                "target_table": "LOCATION_ADDRESS",
                "target_column": "ADDRESS_LINE",
                "required": False,
            },
        ]
    )
    payload["relationship_rules"] = [
        {
            "rule_key": "dynamic_location_address_parent",
            "parent_sheet_code": "LOCATIONS",
            "parent_field_key": "location_gid",
            "child_sheet_code": "LOCATION_ADDRESSES",
            "child_field_key": "address_location_gid",
            "severity": "ERROR",
        }
    ]
    return payload


def dynamic_operational_locations_payload(code: str = "LOCATIONS_OPERATIONAL_QA") -> dict[str, object]:
    return {
        "code": code,
        "name": "Locations Operational QA",
        "catalog_macro_object_code": "LOCATION",
        "data_category": "MASTER_DATA",
        "target_tables": [
            {"table_name": "LOCATION", "sequence": 10, "required": True},
            {"table_name": "LOCATION_ADDRESS", "sequence": 20, "required": False},
            {"table_name": "LOCATION_CAPACITY", "sequence": 30, "required": False},
            {"table_name": "LOCATION_ACTIVITY_TIME_DEF", "sequence": 40, "required": False},
            {"table_name": "LOCATION_LOAD_UNLOAD_POINT", "sequence": 50, "required": False},
            {"table_name": "EQUIPMENT_GROUP_PROFILE", "sequence": 60, "required": False},
            {"table_name": "EQUIPMENT_GROUP_PROFILE_D", "sequence": 70, "required": False},
        ],
        "sheets": [
            {
                "code": "LOCATIONS",
                "name": "Locations",
                "sequence": 10,
                "field_keys": [
                    "location_gid",
                    "location_xid",
                    "location_name",
                    "city",
                    "province_code",
                    "postal_code",
                    "country_code3_gid",
                    "time_zone_gid",
                    "lat",
                    "lon",
                    "location_equipment_group_profile_gid",
                    "appointment_activity_type",
                ],
            },
            {
                "code": "LOCATION_ADDRESSES",
                "name": "Location Addresses",
                "sequence": 20,
                "field_keys": [
                    "address_location_gid",
                    "address_line_sequence",
                    "address_line",
                ],
            },
            {
                "code": "LOCATION_CAPACITIES",
                "name": "Location Capacities",
                "sequence": 30,
                "field_keys": [
                    "capacity_location_gid",
                    "location_capacity_gid",
                    "location_capacity_xid",
                    "capacity_calendar_gid",
                ],
            },
            {
                "code": "LOCATION_ACTIVITY_TIMES",
                "name": "Location Activity Times",
                "sequence": 40,
                "field_keys": [
                    "activity_location_gid",
                    "location_role_gid",
                    "activity_time_def_gid",
                ],
            },
            {
                "code": "LOCATION_DOCKS",
                "name": "Location Docks",
                "sequence": 50,
                "field_keys": [
                    "dock_location_gid",
                    "load_unload_point",
                    "dock_equipment_group_profile_gid",
                    "is_load",
                    "load_sequence",
                    "is_unload",
                    "unload_sequence",
                ],
            },
            {
                "code": "EQUIPMENT_PROFILES",
                "name": "Equipment Profiles",
                "sequence": 60,
                "field_keys": [
                    "equipment_group_profile_gid",
                    "equipment_group_profile_xid",
                    "equipment_group_profile_name",
                ],
            },
            {
                "code": "EQUIPMENT_PROFILE_DETAILS",
                "name": "Equipment Profile Details",
                "sequence": 70,
                "field_keys": [
                    "detail_equipment_group_profile_gid",
                    "equipment_group_gid",
                ],
            },
        ],
        "fields": [
            {"field_key": "location_gid", "label": "Location GID", "data_type": "string", "required": True, "sheet_code": "LOCATIONS"},
            {"field_key": "location_xid", "label": "Location XID", "data_type": "string", "required": True, "sheet_code": "LOCATIONS"},
            {"field_key": "location_name", "label": "Location Name", "data_type": "string", "required": True, "sheet_code": "LOCATIONS"},
            {"field_key": "city", "label": "City", "data_type": "string", "required": False, "sheet_code": "LOCATIONS"},
            {"field_key": "province_code", "label": "Province Code", "data_type": "string", "required": False, "sheet_code": "LOCATIONS"},
            {"field_key": "postal_code", "label": "Postal Code", "data_type": "string", "required": False, "sheet_code": "LOCATIONS"},
            {"field_key": "country_code3_gid", "label": "Country Code", "data_type": "string", "required": False, "sheet_code": "LOCATIONS"},
            {"field_key": "time_zone_gid", "label": "Time Zone", "data_type": "string", "required": False, "sheet_code": "LOCATIONS"},
            {"field_key": "lat", "label": "Latitude", "data_type": "number", "required": False, "sheet_code": "LOCATIONS"},
            {"field_key": "lon", "label": "Longitude", "data_type": "number", "required": False, "sheet_code": "LOCATIONS"},
            {"field_key": "location_equipment_group_profile_gid", "label": "Equipment Restriction Profile", "data_type": "string", "required": False, "sheet_code": "LOCATIONS"},
            {"field_key": "appointment_activity_type", "label": "Appointment Activity Type", "data_type": "string", "required": False, "sheet_code": "LOCATIONS"},
            {"field_key": "address_location_gid", "label": "Address Location GID", "data_type": "string", "required": True, "sheet_code": "LOCATION_ADDRESSES"},
            {"field_key": "address_line_sequence", "label": "Address Line Sequence", "data_type": "number", "required": True, "sheet_code": "LOCATION_ADDRESSES"},
            {"field_key": "address_line", "label": "Address Line", "data_type": "string", "required": False, "sheet_code": "LOCATION_ADDRESSES"},
            {"field_key": "capacity_location_gid", "label": "Capacity Location GID", "data_type": "string", "required": True, "sheet_code": "LOCATION_CAPACITIES"},
            {"field_key": "location_capacity_gid", "label": "Location Capacity GID", "data_type": "string", "required": True, "sheet_code": "LOCATION_CAPACITIES"},
            {"field_key": "location_capacity_xid", "label": "Location Capacity XID", "data_type": "string", "required": True, "sheet_code": "LOCATION_CAPACITIES"},
            {"field_key": "capacity_calendar_gid", "label": "Capacity Calendar GID", "data_type": "string", "required": False, "sheet_code": "LOCATION_CAPACITIES"},
            {"field_key": "activity_location_gid", "label": "Activity Location GID", "data_type": "string", "required": True, "sheet_code": "LOCATION_ACTIVITY_TIMES"},
            {"field_key": "location_role_gid", "label": "Location Role GID", "data_type": "string", "required": True, "sheet_code": "LOCATION_ACTIVITY_TIMES"},
            {"field_key": "activity_time_def_gid", "label": "Activity Time Definition GID", "data_type": "string", "required": True, "sheet_code": "LOCATION_ACTIVITY_TIMES"},
            {"field_key": "dock_location_gid", "label": "Dock Location GID", "data_type": "string", "required": True, "sheet_code": "LOCATION_DOCKS"},
            {"field_key": "load_unload_point", "label": "Dock Code", "data_type": "string", "required": True, "sheet_code": "LOCATION_DOCKS"},
            {"field_key": "dock_equipment_group_profile_gid", "label": "Dock Equipment Profile", "data_type": "string", "required": False, "sheet_code": "LOCATION_DOCKS"},
            {"field_key": "is_load", "label": "Is Load Dock", "data_type": "string", "required": False, "sheet_code": "LOCATION_DOCKS"},
            {"field_key": "load_sequence", "label": "Load Sequence", "data_type": "number", "required": False, "sheet_code": "LOCATION_DOCKS"},
            {"field_key": "is_unload", "label": "Is Unload Dock", "data_type": "string", "required": False, "sheet_code": "LOCATION_DOCKS"},
            {"field_key": "unload_sequence", "label": "Unload Sequence", "data_type": "number", "required": False, "sheet_code": "LOCATION_DOCKS"},
            {"field_key": "equipment_group_profile_xid", "label": "Equipment Profile XID", "data_type": "string", "required": True, "sheet_code": "EQUIPMENT_PROFILES"},
            {"field_key": "equipment_group_profile_gid", "label": "Equipment Profile GID", "data_type": "string", "required": True, "sheet_code": "EQUIPMENT_PROFILES"},
            {"field_key": "equipment_group_profile_name", "label": "Equipment Profile Name", "data_type": "string", "required": False, "sheet_code": "EQUIPMENT_PROFILES"},
            {"field_key": "detail_equipment_group_profile_gid", "label": "Detail Equipment Profile GID", "data_type": "string", "required": True, "sheet_code": "EQUIPMENT_PROFILE_DETAILS"},
            {"field_key": "equipment_group_gid", "label": "Equipment Group GID", "data_type": "string", "required": True, "sheet_code": "EQUIPMENT_PROFILE_DETAILS"},
        ],
        "mappings": [
            {"mapping_key": "location_gid", "source_type": "USER_FIELD", "source_field_key": "location_gid", "target_table": "LOCATION", "target_column": "LOCATION_GID", "required": True},
            {"mapping_key": "location_xid", "source_type": "USER_FIELD", "source_field_key": "location_xid", "target_table": "LOCATION", "target_column": "LOCATION_XID", "required": True},
            {"mapping_key": "location_name", "source_type": "USER_FIELD", "source_field_key": "location_name", "target_table": "LOCATION", "target_column": "LOCATION_NAME", "required": True},
            {"mapping_key": "location_city", "source_type": "USER_FIELD", "source_field_key": "city", "target_table": "LOCATION", "target_column": "CITY", "required": False},
            {"mapping_key": "location_province", "source_type": "USER_FIELD", "source_field_key": "province_code", "target_table": "LOCATION", "target_column": "PROVINCE_CODE", "required": False},
            {"mapping_key": "location_postal", "source_type": "USER_FIELD", "source_field_key": "postal_code", "target_table": "LOCATION", "target_column": "POSTAL_CODE", "required": False},
            {"mapping_key": "location_country", "source_type": "USER_FIELD", "source_field_key": "country_code3_gid", "target_table": "LOCATION", "target_column": "COUNTRY_CODE3_GID", "required": False},
            {"mapping_key": "location_timezone", "source_type": "USER_FIELD", "source_field_key": "time_zone_gid", "target_table": "LOCATION", "target_column": "TIME_ZONE_GID", "required": False},
            {"mapping_key": "location_lat", "source_type": "USER_FIELD", "source_field_key": "lat", "target_table": "LOCATION", "target_column": "LAT", "required": False},
            {"mapping_key": "location_lon", "source_type": "USER_FIELD", "source_field_key": "lon", "target_table": "LOCATION", "target_column": "LON", "required": False},
            {"mapping_key": "location_equipment_profile", "source_type": "USER_FIELD", "source_field_key": "location_equipment_group_profile_gid", "target_table": "LOCATION", "target_column": "SHUTTLE_LOT_EQ_GRP_PROFILE_GID", "required": False},
            {"mapping_key": "location_activity_type", "source_type": "USER_FIELD", "source_field_key": "appointment_activity_type", "target_table": "LOCATION", "target_column": "APPOINTMENT_ACTIVITY_TYPE", "required": False},
            {"mapping_key": "address_location", "source_type": "USER_FIELD", "source_field_key": "address_location_gid", "target_table": "LOCATION_ADDRESS", "target_column": "LOCATION_GID", "required": True},
            {"mapping_key": "address_sequence", "source_type": "USER_FIELD", "source_field_key": "address_line_sequence", "target_table": "LOCATION_ADDRESS", "target_column": "LINE_SEQUENCE", "required": True},
            {"mapping_key": "address_line", "source_type": "USER_FIELD", "source_field_key": "address_line", "target_table": "LOCATION_ADDRESS", "target_column": "ADDRESS_LINE", "required": False},
            {"mapping_key": "capacity_gid", "source_type": "USER_FIELD", "source_field_key": "location_capacity_gid", "target_table": "LOCATION_CAPACITY", "target_column": "LOCATION_CAPACITY_GID", "required": True},
            {"mapping_key": "capacity_xid", "source_type": "USER_FIELD", "source_field_key": "location_capacity_xid", "target_table": "LOCATION_CAPACITY", "target_column": "LOCATION_CAPACITY_XID", "required": True},
            {"mapping_key": "capacity_calendar", "source_type": "USER_FIELD", "source_field_key": "capacity_calendar_gid", "target_table": "LOCATION_CAPACITY", "target_column": "CALENDAR_GID", "required": False},
            {"mapping_key": "activity_location", "source_type": "USER_FIELD", "source_field_key": "activity_location_gid", "target_table": "LOCATION_ACTIVITY_TIME_DEF", "target_column": "LOCATION_GID", "required": True},
            {"mapping_key": "activity_role", "source_type": "USER_FIELD", "source_field_key": "location_role_gid", "target_table": "LOCATION_ACTIVITY_TIME_DEF", "target_column": "LOCATION_ROLE_GID", "required": True},
            {"mapping_key": "activity_time_def", "source_type": "USER_FIELD", "source_field_key": "activity_time_def_gid", "target_table": "LOCATION_ACTIVITY_TIME_DEF", "target_column": "ACTIVITY_TIME_DEF_GID", "required": True},
            {"mapping_key": "dock_location", "source_type": "USER_FIELD", "source_field_key": "dock_location_gid", "target_table": "LOCATION_LOAD_UNLOAD_POINT", "target_column": "LOCATION_GID", "required": True},
            {"mapping_key": "dock_point", "source_type": "USER_FIELD", "source_field_key": "load_unload_point", "target_table": "LOCATION_LOAD_UNLOAD_POINT", "target_column": "LOAD_UNLOAD_POINT", "required": True},
            {"mapping_key": "dock_profile", "source_type": "USER_FIELD", "source_field_key": "dock_equipment_group_profile_gid", "target_table": "LOCATION_LOAD_UNLOAD_POINT", "target_column": "EQUIPMENT_GROUP_PROFILE_GID", "required": False},
            {"mapping_key": "dock_is_load", "source_type": "USER_FIELD", "source_field_key": "is_load", "target_table": "LOCATION_LOAD_UNLOAD_POINT", "target_column": "IS_LOAD", "required": False},
            {"mapping_key": "dock_load_sequence", "source_type": "USER_FIELD", "source_field_key": "load_sequence", "target_table": "LOCATION_LOAD_UNLOAD_POINT", "target_column": "LOAD_SEQUENCE", "required": False},
            {"mapping_key": "dock_is_unload", "source_type": "USER_FIELD", "source_field_key": "is_unload", "target_table": "LOCATION_LOAD_UNLOAD_POINT", "target_column": "IS_UNLOAD", "required": False},
            {"mapping_key": "dock_unload_sequence", "source_type": "USER_FIELD", "source_field_key": "unload_sequence", "target_table": "LOCATION_LOAD_UNLOAD_POINT", "target_column": "UNLOAD_SEQUENCE", "required": False},
            {"mapping_key": "profile_gid", "source_type": "USER_FIELD", "source_field_key": "equipment_group_profile_gid", "target_table": "EQUIPMENT_GROUP_PROFILE", "target_column": "EQUIPMENT_GROUP_PROFILE_GID", "required": True},
            {"mapping_key": "profile_xid", "source_type": "USER_FIELD", "source_field_key": "equipment_group_profile_xid", "target_table": "EQUIPMENT_GROUP_PROFILE", "target_column": "EQUIPMENT_GROUP_PROFILE_XID", "required": True},
            {"mapping_key": "profile_name", "source_type": "USER_FIELD", "source_field_key": "equipment_group_profile_name", "target_table": "EQUIPMENT_GROUP_PROFILE", "target_column": "EQUIPMENT_GROUP_PROFILE_NAME", "required": False},
            {"mapping_key": "profile_detail_gid", "source_type": "USER_FIELD", "source_field_key": "detail_equipment_group_profile_gid", "target_table": "EQUIPMENT_GROUP_PROFILE_D", "target_column": "EQUIPMENT_GROUP_PROFILE_GID", "required": True},
            {"mapping_key": "profile_detail_equipment_group", "source_type": "USER_FIELD", "source_field_key": "equipment_group_gid", "target_table": "EQUIPMENT_GROUP_PROFILE_D", "target_column": "EQUIPMENT_GROUP_GID", "required": True},
        ],
        "relationship_rules": [
            {"rule_key": "location_address_parent", "parent_sheet_code": "LOCATIONS", "parent_field_key": "location_gid", "child_sheet_code": "LOCATION_ADDRESSES", "child_field_key": "address_location_gid", "severity": "ERROR"},
            {"rule_key": "location_capacity_parent", "parent_sheet_code": "LOCATIONS", "parent_field_key": "location_gid", "child_sheet_code": "LOCATION_CAPACITIES", "child_field_key": "capacity_location_gid", "severity": "ERROR"},
            {"rule_key": "location_activity_parent", "parent_sheet_code": "LOCATIONS", "parent_field_key": "location_gid", "child_sheet_code": "LOCATION_ACTIVITY_TIMES", "child_field_key": "activity_location_gid", "severity": "ERROR"},
            {"rule_key": "location_dock_parent", "parent_sheet_code": "LOCATIONS", "parent_field_key": "location_gid", "child_sheet_code": "LOCATION_DOCKS", "child_field_key": "dock_location_gid", "severity": "ERROR"},
            {"rule_key": "location_profile_parent", "parent_sheet_code": "EQUIPMENT_PROFILES", "parent_field_key": "equipment_group_profile_gid", "child_sheet_code": "LOCATIONS", "child_field_key": "location_equipment_group_profile_gid", "severity": "ERROR"},
            {"rule_key": "profile_detail_parent", "parent_sheet_code": "EQUIPMENT_PROFILES", "parent_field_key": "equipment_group_profile_gid", "child_sheet_code": "EQUIPMENT_PROFILE_DETAILS", "child_field_key": "detail_equipment_group_profile_gid", "severity": "ERROR"},
        ],
        "documentation_refs": [
            {"source_type": "DATA_DICTIONARY", "scope": "LOCATION, LOCATION_ADDRESS, LOCATION_CAPACITY, LOCATION_ACTIVITY_TIME_DEF, LOCATION_LOAD_UNLOAD_POINT, EQUIPMENT_GROUP_PROFILE, EQUIPMENT_GROUP_PROFILE_D", "note": "Validated against local OTM Data Dictionary 26B."},
            {"source_type": "ORACLE_OFFICIAL", "scope": "Equipment Group Profile, Activity Type Capacity, Activity Time Definition", "note": "Oracle Help confirms equipment profile restrictions, location capacity, and activity time concepts."},
        ],
    }


def dynamic_item_packaging_payload(code: str = "ITEM_PACKAGING_OPERATIONAL_QA") -> dict[str, object]:
    return {
        "code": code,
        "name": "Item Packaging Operational QA",
        "catalog_macro_object_code": "ITEM",
        "data_category": "MASTER_DATA",
        "target_tables": [
            {"table_name": "ITEM", "sequence": 10, "required": True},
            {"table_name": "SHIP_UNIT_SPEC", "sequence": 20, "required": True},
            {"table_name": "PACKAGED_ITEM", "sequence": 30, "required": True},
            {"table_name": "TI_HI", "sequence": 40, "required": False},
        ],
        "sheets": [
            {"code": "ITEMS", "name": "Items", "sequence": 10, "field_keys": ["item_gid", "item_xid", "item_name", "item_type_gid", "unit_of_measure"]},
            {"code": "SHIP_UNIT_SPECS", "name": "Ship Unit Specs", "sequence": 20, "field_keys": ["ship_unit_spec_gid", "ship_unit_spec_xid", "ship_unit_spec_name", "unit_type", "length", "length_uom_code", "width", "width_uom_code", "height", "height_uom_code", "tare_weight", "tare_weight_uom_code", "effective_volume", "effective_volume_uom_code", "is_in_on_max"]},
            {"code": "PACKAGED_ITEMS", "name": "Packaged Items", "sequence": 30, "field_keys": ["packaged_item_gid", "packaged_item_xid", "package_item_gid", "packaging_unit_gid", "package_ship_unit_weight", "package_ship_unit_weight_uom", "package_su_volume", "package_su_volume_uom_code", "package_su_length", "package_su_length_uom_code", "package_su_width", "package_su_width_uom_code", "package_su_height", "package_su_height_uom_code"]},
            {"code": "TI_HI", "name": "TiHi", "sequence": 40, "field_keys": ["tihi_sequence_no", "tihi_packaged_item_gid", "tihi_packaging_unit_gid", "transport_handling_unit_gid", "num_layers", "quantity_per_layer"]},
        ],
        "fields": [
            {"field_key": "item_gid", "label": "Item GID", "data_type": "string", "required": True, "sheet_code": "ITEMS"},
            {"field_key": "item_xid", "label": "Item XID", "data_type": "string", "required": True, "sheet_code": "ITEMS"},
            {"field_key": "item_name", "label": "Item Name", "data_type": "string", "required": False, "sheet_code": "ITEMS"},
            {"field_key": "item_type_gid", "label": "Item Type GID", "data_type": "string", "required": False, "sheet_code": "ITEMS"},
            {"field_key": "unit_of_measure", "label": "Unit of Measure", "data_type": "string", "required": False, "sheet_code": "ITEMS"},
            {"field_key": "ship_unit_spec_gid", "label": "Ship Unit Spec GID", "data_type": "string", "required": True, "sheet_code": "SHIP_UNIT_SPECS"},
            {"field_key": "ship_unit_spec_xid", "label": "Ship Unit Spec XID", "data_type": "string", "required": True, "sheet_code": "SHIP_UNIT_SPECS"},
            {"field_key": "ship_unit_spec_name", "label": "Ship Unit Spec Name", "data_type": "string", "required": False, "sheet_code": "SHIP_UNIT_SPECS"},
            {"field_key": "unit_type", "label": "Unit Type", "data_type": "string", "required": True, "sheet_code": "SHIP_UNIT_SPECS"},
            {"field_key": "length", "label": "Length", "data_type": "number", "required": False, "sheet_code": "SHIP_UNIT_SPECS"},
            {"field_key": "length_uom_code", "label": "Length UOM", "data_type": "string", "required": False, "sheet_code": "SHIP_UNIT_SPECS"},
            {"field_key": "width", "label": "Width", "data_type": "number", "required": False, "sheet_code": "SHIP_UNIT_SPECS"},
            {"field_key": "width_uom_code", "label": "Width UOM", "data_type": "string", "required": False, "sheet_code": "SHIP_UNIT_SPECS"},
            {"field_key": "height", "label": "Height", "data_type": "number", "required": False, "sheet_code": "SHIP_UNIT_SPECS"},
            {"field_key": "height_uom_code", "label": "Height UOM", "data_type": "string", "required": False, "sheet_code": "SHIP_UNIT_SPECS"},
            {"field_key": "tare_weight", "label": "Tare Weight", "data_type": "number", "required": False, "sheet_code": "SHIP_UNIT_SPECS"},
            {"field_key": "tare_weight_uom_code", "label": "Tare Weight UOM", "data_type": "string", "required": False, "sheet_code": "SHIP_UNIT_SPECS"},
            {"field_key": "effective_volume", "label": "Effective Volume", "data_type": "number", "required": False, "sheet_code": "SHIP_UNIT_SPECS"},
            {"field_key": "effective_volume_uom_code", "label": "Effective Volume UOM", "data_type": "string", "required": False, "sheet_code": "SHIP_UNIT_SPECS"},
            {"field_key": "is_in_on_max", "label": "Inside/On Max", "data_type": "string", "required": False, "sheet_code": "SHIP_UNIT_SPECS"},
            {"field_key": "packaged_item_gid", "label": "Packaged Item GID", "data_type": "string", "required": True, "sheet_code": "PACKAGED_ITEMS"},
            {"field_key": "packaged_item_xid", "label": "Packaged Item XID", "data_type": "string", "required": True, "sheet_code": "PACKAGED_ITEMS"},
            {"field_key": "package_item_gid", "label": "Item GID", "data_type": "string", "required": True, "sheet_code": "PACKAGED_ITEMS"},
            {"field_key": "packaging_unit_gid", "label": "Packaging Unit GID", "data_type": "string", "required": True, "sheet_code": "PACKAGED_ITEMS"},
            {"field_key": "package_ship_unit_weight", "label": "Package Weight", "data_type": "number", "required": False, "sheet_code": "PACKAGED_ITEMS"},
            {"field_key": "package_ship_unit_weight_uom", "label": "Package Weight UOM", "data_type": "string", "required": False, "sheet_code": "PACKAGED_ITEMS"},
            {"field_key": "package_su_volume", "label": "Package Volume", "data_type": "number", "required": False, "sheet_code": "PACKAGED_ITEMS"},
            {"field_key": "package_su_volume_uom_code", "label": "Package Volume UOM", "data_type": "string", "required": False, "sheet_code": "PACKAGED_ITEMS"},
            {"field_key": "package_su_length", "label": "Package Length", "data_type": "number", "required": False, "sheet_code": "PACKAGED_ITEMS"},
            {"field_key": "package_su_length_uom_code", "label": "Package Length UOM", "data_type": "string", "required": False, "sheet_code": "PACKAGED_ITEMS"},
            {"field_key": "package_su_width", "label": "Package Width", "data_type": "number", "required": False, "sheet_code": "PACKAGED_ITEMS"},
            {"field_key": "package_su_width_uom_code", "label": "Package Width UOM", "data_type": "string", "required": False, "sheet_code": "PACKAGED_ITEMS"},
            {"field_key": "package_su_height", "label": "Package Height", "data_type": "number", "required": False, "sheet_code": "PACKAGED_ITEMS"},
            {"field_key": "package_su_height_uom_code", "label": "Package Height UOM", "data_type": "string", "required": False, "sheet_code": "PACKAGED_ITEMS"},
            {"field_key": "tihi_sequence_no", "label": "TiHi Sequence", "data_type": "number", "required": True, "sheet_code": "TI_HI"},
            {"field_key": "tihi_packaged_item_gid", "label": "Packaged Item GID", "data_type": "string", "required": True, "sheet_code": "TI_HI"},
            {"field_key": "tihi_packaging_unit_gid", "label": "Packaging Unit GID", "data_type": "string", "required": True, "sheet_code": "TI_HI"},
            {"field_key": "transport_handling_unit_gid", "label": "Transport Handling Unit GID", "data_type": "string", "required": True, "sheet_code": "TI_HI"},
            {"field_key": "num_layers", "label": "Number of Layers", "data_type": "number", "required": True, "sheet_code": "TI_HI"},
            {"field_key": "quantity_per_layer", "label": "Quantity per Layer", "data_type": "number", "required": True, "sheet_code": "TI_HI"},
        ],
        "mappings": [
            {"mapping_key": "item_gid", "source_type": "USER_FIELD", "source_field_key": "item_gid", "target_table": "ITEM", "target_column": "ITEM_GID", "required": True},
            {"mapping_key": "item_xid", "source_type": "USER_FIELD", "source_field_key": "item_xid", "target_table": "ITEM", "target_column": "ITEM_XID", "required": True},
            {"mapping_key": "item_name", "source_type": "USER_FIELD", "source_field_key": "item_name", "target_table": "ITEM", "target_column": "ITEM_NAME", "required": False},
            {"mapping_key": "item_type", "source_type": "USER_FIELD", "source_field_key": "item_type_gid", "target_table": "ITEM", "target_column": "ITEM_TYPE_GID", "required": False},
            {"mapping_key": "item_uom", "source_type": "USER_FIELD", "source_field_key": "unit_of_measure", "target_table": "ITEM", "target_column": "UNIT_OF_MEASURE", "required": False},
            *[
                {"mapping_key": f"sus_{field}", "source_type": "USER_FIELD", "source_field_key": field.lower(), "target_table": "SHIP_UNIT_SPEC", "target_column": field, "required": field in {"SHIP_UNIT_SPEC_GID", "SHIP_UNIT_SPEC_XID", "UNIT_TYPE"}}
                for field in [
                    "SHIP_UNIT_SPEC_GID", "SHIP_UNIT_SPEC_XID", "SHIP_UNIT_SPEC_NAME", "UNIT_TYPE", "LENGTH", "LENGTH_UOM_CODE", "WIDTH", "WIDTH_UOM_CODE", "HEIGHT", "HEIGHT_UOM_CODE", "TARE_WEIGHT", "TARE_WEIGHT_UOM_CODE", "EFFECTIVE_VOLUME", "EFFECTIVE_VOLUME_UOM_CODE", "IS_IN_ON_MAX"
                ]
            ],
            {"mapping_key": "packaged_item_gid", "source_type": "USER_FIELD", "source_field_key": "packaged_item_gid", "target_table": "PACKAGED_ITEM", "target_column": "PACKAGED_ITEM_GID", "required": True},
            {"mapping_key": "packaged_item_xid", "source_type": "USER_FIELD", "source_field_key": "packaged_item_xid", "target_table": "PACKAGED_ITEM", "target_column": "PACKAGED_ITEM_XID", "required": True},
            {"mapping_key": "packaged_item_item", "source_type": "USER_FIELD", "source_field_key": "package_item_gid", "target_table": "PACKAGED_ITEM", "target_column": "ITEM_GID", "required": True},
            {"mapping_key": "packaged_item_packaging_unit", "source_type": "USER_FIELD", "source_field_key": "packaging_unit_gid", "target_table": "PACKAGED_ITEM", "target_column": "PACKAGING_UNIT_GID", "required": True},
            *[
                {"mapping_key": f"packaged_item_{column.lower()}", "source_type": "USER_FIELD", "source_field_key": column.lower(), "target_table": "PACKAGED_ITEM", "target_column": column, "required": False}
                for column in [
                    "PACKAGE_SHIP_UNIT_WEIGHT", "PACKAGE_SHIP_UNIT_WEIGHT_UOM", "PACKAGE_SU_VOLUME", "PACKAGE_SU_VOLUME_UOM_CODE", "PACKAGE_SU_LENGTH", "PACKAGE_SU_LENGTH_UOM_CODE", "PACKAGE_SU_WIDTH", "PACKAGE_SU_WIDTH_UOM_CODE", "PACKAGE_SU_HEIGHT", "PACKAGE_SU_HEIGHT_UOM_CODE"
                ]
            ],
            {"mapping_key": "tihi_sequence", "source_type": "USER_FIELD", "source_field_key": "tihi_sequence_no", "target_table": "TI_HI", "target_column": "SEQUENCE_NO", "required": True},
            {"mapping_key": "tihi_packaged_item", "source_type": "USER_FIELD", "source_field_key": "tihi_packaged_item_gid", "target_table": "TI_HI", "target_column": "PACKAGED_ITEM_GID", "required": True},
            {"mapping_key": "tihi_packaging_unit", "source_type": "USER_FIELD", "source_field_key": "tihi_packaging_unit_gid", "target_table": "TI_HI", "target_column": "PACKAGING_UNIT_GID", "required": True},
            {"mapping_key": "tihi_thu", "source_type": "USER_FIELD", "source_field_key": "transport_handling_unit_gid", "target_table": "TI_HI", "target_column": "TRANSPORT_HANDLING_UNIT_GID", "required": True},
            {"mapping_key": "tihi_layers", "source_type": "USER_FIELD", "source_field_key": "num_layers", "target_table": "TI_HI", "target_column": "NUM_LAYERS", "required": True},
            {"mapping_key": "tihi_qty_layer", "source_type": "USER_FIELD", "source_field_key": "quantity_per_layer", "target_table": "TI_HI", "target_column": "QUANTITY_PER_LAYER", "required": True},
        ],
        "relationship_rules": [
            {"rule_key": "packaged_item_item_parent", "parent_sheet_code": "ITEMS", "parent_field_key": "item_gid", "child_sheet_code": "PACKAGED_ITEMS", "child_field_key": "package_item_gid", "severity": "ERROR"},
            {"rule_key": "packaged_item_packaging_unit_parent", "parent_sheet_code": "SHIP_UNIT_SPECS", "parent_field_key": "ship_unit_spec_gid", "child_sheet_code": "PACKAGED_ITEMS", "child_field_key": "packaging_unit_gid", "severity": "ERROR"},
            {"rule_key": "tihi_packaged_item_parent", "parent_sheet_code": "PACKAGED_ITEMS", "parent_field_key": "packaged_item_gid", "child_sheet_code": "TI_HI", "child_field_key": "tihi_packaged_item_gid", "severity": "ERROR"},
            {"rule_key": "tihi_packaging_unit_parent", "parent_sheet_code": "SHIP_UNIT_SPECS", "parent_field_key": "ship_unit_spec_gid", "child_sheet_code": "TI_HI", "child_field_key": "tihi_packaging_unit_gid", "severity": "ERROR"},
            {"rule_key": "tihi_thu_parent", "parent_sheet_code": "SHIP_UNIT_SPECS", "parent_field_key": "ship_unit_spec_gid", "child_sheet_code": "TI_HI", "child_field_key": "transport_handling_unit_gid", "severity": "ERROR"},
        ],
        "documentation_refs": [
            {"source_type": "DATA_DICTIONARY", "scope": "ITEM, SHIP_UNIT_SPEC, PACKAGED_ITEM, TI_HI", "note": "Validated against local OTM Data Dictionary 26B."},
            {"source_type": "ORACLE_OFFICIAL", "scope": "Packaged Item, TiHi, Ship Unit Calculation Logic", "note": "Oracle Help confirms the Item > Packaged Item > Ship Unit / TiHi hierarchy."},
        ],
    }


def test_master_data_dynamic_template_draft_create_exposes_v2_definition(client, admin_header):
    response = client.post(
        "/api/v1/modules/master-data/templates/drafts",
        json=dynamic_locations_template_payload(),
        headers=admin_header,
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["code"] == "LOCATIONS_DYNAMIC"
    assert payload["status"] == "DRAFT"
    assert payload["definition"]["schema_version"] == "master-data-template-definition/v2"
    assert payload["definition"]["mappings"][1]["target_column"] == "LOCATION_XID"


def test_master_data_dynamic_template_draft_patch_updates_definition(client, admin_header):
    create = client.post(
        "/api/v1/modules/master-data/templates/drafts",
        json=dynamic_locations_template_payload("LOCATIONS_DYNAMIC_PATCH"),
        headers=admin_header,
    )
    assert create.status_code == 200
    payload = dynamic_locations_template_payload("LOCATIONS_DYNAMIC_PATCH")
    payload["name"] = "Locations Dynamic Patched"
    payload["fields"][1]["label"] = "Friendly Location Name"

    response = client.patch(
        "/api/v1/modules/master-data/templates/LOCATIONS_DYNAMIC_PATCH/draft",
        json=payload,
        headers=admin_header,
    )

    assert response.status_code == 200
    patched = response.json()
    assert patched["name"] == "Locations Dynamic Patched"
    assert patched["status"] == "DRAFT"
    assert patched["definition"]["fields"][1]["label"] == "Friendly Location Name"


def test_master_data_dynamic_template_draft_patch_rejects_published_template(client, admin_header):
    create = client.post(
        "/api/v1/modules/master-data/templates/drafts",
        json=dynamic_locations_template_payload("LOCATIONS_DYNAMIC_PATCH_PUBLISHED"),
        headers=admin_header,
    )
    assert create.status_code == 200
    publish = client.post(
        "/api/v1/modules/master-data/templates/LOCATIONS_DYNAMIC_PATCH_PUBLISHED/publish",
        headers=admin_header,
    )
    assert publish.status_code == 200

    response = client.patch(
        "/api/v1/modules/master-data/templates/LOCATIONS_DYNAMIC_PATCH_PUBLISHED/draft",
        json=dynamic_locations_template_payload("LOCATIONS_DYNAMIC_PATCH_PUBLISHED"),
        headers=admin_header,
    )

    assert response.status_code == 409
    assert response.json()["code"] == "MASTER_DATA_TEMPLATE_DRAFT_NOT_UPDATABLE"


def test_master_data_dynamic_template_validation_rejects_invalid_target_column(client, admin_header):
    payload = dynamic_locations_template_payload("LOCATIONS_BAD_COLUMN")
    payload["mappings"][0]["target_column"] = "NOT_A_REAL_COLUMN"

    create = client.post(
        "/api/v1/modules/master-data/templates/drafts",
        json=payload,
        headers=admin_header,
    )
    assert create.status_code == 200

    response = client.post(
        "/api/v1/modules/master-data/templates/LOCATIONS_BAD_COLUMN/validate-definition",
        headers=admin_header,
    )

    assert response.status_code == 200
    validation = response.json()
    assert validation["valid"] is False
    assert validation["severity"] == "ERROR"
    assert validation["issues"][0]["code"] == "MASTER_DATA_TEMPLATE_MAPPING_COLUMN_INVALID"
    assert validation["issues"][0]["target_column"] == "NOT_A_REAL_COLUMN"


def test_master_data_dynamic_template_validation_requires_documentation_refs(client, admin_header):
    payload = dynamic_locations_template_payload("LOCATIONS_NO_DOC_REFS")
    payload["documentation_refs"] = []

    create = client.post(
        "/api/v1/modules/master-data/templates/drafts",
        json=payload,
        headers=admin_header,
    )
    assert create.status_code == 200

    response = client.post(
        "/api/v1/modules/master-data/templates/LOCATIONS_NO_DOC_REFS/validate-definition",
        headers=admin_header,
    )

    assert response.status_code == 200
    validation = response.json()
    assert validation["valid"] is False
    assert validation["issues"][0]["code"] == "MASTER_DATA_TEMPLATE_DOCUMENTATION_REF_REQUIRED"


def test_master_data_dynamic_template_publish_validates_one_to_many_and_fixed_value(
    client,
    admin_header,
):
    create = client.post(
        "/api/v1/modules/master-data/templates/drafts",
        json=dynamic_locations_template_payload("LOCATIONS_DYNAMIC_PUBLISH"),
        headers=admin_header,
    )
    assert create.status_code == 200

    response = client.post(
        "/api/v1/modules/master-data/templates/LOCATIONS_DYNAMIC_PUBLISH/publish",
        headers=admin_header,
    )

    assert response.status_code == 200
    published = response.json()
    assert published["status"] == "PUBLISHED"
    assert published["validation"]["valid"] is True
    assert published["validation"]["summary"]["mapping_count"] == 4
    assert published["definition"]["mappings"][1]["source_field_key"] == "location_gid"
    assert published["definition"]["mappings"][3]["source_type"] == "FIXED_VALUE"


def test_master_data_dynamic_template_create_next_version_as_draft(client, admin_header):
    client.post(
        "/api/v1/modules/master-data/templates/drafts",
        json=dynamic_locations_template_payload("LOCATIONS_DYNAMIC_VERSION_SOURCE"),
        headers=admin_header,
    )
    publish = client.post(
        "/api/v1/modules/master-data/templates/LOCATIONS_DYNAMIC_VERSION_SOURCE/publish",
        headers=admin_header,
    )
    assert publish.status_code == 200

    response = client.post(
        "/api/v1/modules/master-data/templates/LOCATIONS_DYNAMIC_VERSION_SOURCE/versions",
        json={"new_code": "LOCATIONS_DYNAMIC_VERSION_2"},
        headers=admin_header,
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["code"] == "LOCATIONS_DYNAMIC_VERSION_2"
    assert payload["status"] == "DRAFT"
    assert payload["version"] == 2
    assert payload["definition"]["template"]["status"] == "DRAFT"
    assert payload["definition"]["template"]["version"] == 2
    assert payload["definition"]["mappings"][1]["target_column"] == "LOCATION_XID"


def test_master_data_dynamic_template_runtime_rejects_draft_template(client, admin_header):
    create = client.post(
        "/api/v1/modules/master-data/templates/drafts",
        json=dynamic_locations_template_payload("LOCATIONS_DRAFT_RUNTIME"),
        headers=admin_header,
    )
    assert create.status_code == 200

    response = client.post(
        "/api/v1/modules/master-data/templates/LOCATIONS_DRAFT_RUNTIME/build-workbook",
        headers=admin_header,
    )

    assert response.status_code == 409
    assert response.json()["code"] == "MASTER_DATA_TEMPLATE_NOT_PUBLISHED"


def test_master_data_dynamic_template_runtime_maps_one_to_many_and_fixed_value(
    client,
    admin_header,
):
    client.post(
        "/api/v1/modules/master-data/templates/drafts",
        json=dynamic_locations_template_payload("LOCATIONS_DYNAMIC_RUNTIME"),
        headers=admin_header,
    )
    publish = client.post(
        "/api/v1/modules/master-data/templates/LOCATIONS_DYNAMIC_RUNTIME/publish",
        headers=admin_header,
    )
    assert publish.status_code == 200

    workbook = Workbook()
    locations = workbook.active
    locations.title = "LOCATIONS"
    locations.append(["Location ID", "Location Name"])
    locations.append(["SYN.LOCATION_001", "Synthetic Location"])
    workbook_bytes = BytesIO()
    workbook.save(workbook_bytes)
    workbook_bytes.seek(0)
    batch_response = client.post(
        "/api/v1/modules/master-data/templates/LOCATIONS_DYNAMIC_RUNTIME/batches",
        headers=admin_header,
        files={
            "file": (
                "locations_dynamic_runtime.xlsx",
                workbook_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert batch_response.status_code == 200
    batch_id = batch_response.json()["batch_id"]

    response = client.post(f"/api/v1/modules/master-data/batches/{batch_id}/map", headers=admin_header)

    assert response.status_code == 200
    payload = response.json()["records"][0]["payload"]
    assert payload["LOCATION_GID"] == "SYN.LOCATION_001"
    assert payload["LOCATION_XID"] == "SYN.LOCATION_001"
    assert payload["LOCATION_NAME"] == "Synthetic Location"
    assert payload["COUNTRY_CODE3_GID"] == "USA"


def test_master_data_dynamic_template_runtime_applies_default_value_when_source_is_blank(
    client,
    admin_header,
):
    payload = dynamic_locations_template_payload("LOCATIONS_DYNAMIC_DEFAULT")
    payload["mappings"].append(
        {
            "mapping_key": "city_default_from_blank_location_name",
            "source_type": "DEFAULT_VALUE",
            "source_field_key": "location_name",
            "default_value": "Fallback City",
            "target_table": "LOCATION",
            "target_column": "CITY",
            "required": False,
        }
    )
    client.post(
        "/api/v1/modules/master-data/templates/drafts",
        json=payload,
        headers=admin_header,
    )
    publish = client.post(
        "/api/v1/modules/master-data/templates/LOCATIONS_DYNAMIC_DEFAULT/publish",
        headers=admin_header,
    )
    assert publish.status_code == 200
    workbook = Workbook()
    locations = workbook.active
    locations.title = "LOCATIONS"
    locations.append(["Location ID", "Location Name"])
    locations.append(["SYN.LOCATION_001", ""])
    workbook_bytes = BytesIO()
    workbook.save(workbook_bytes)
    workbook_bytes.seek(0)
    batch_response = client.post(
        "/api/v1/modules/master-data/templates/LOCATIONS_DYNAMIC_DEFAULT/batches",
        headers=admin_header,
        files={
            "file": (
                "locations_dynamic_default.xlsx",
                workbook_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert batch_response.status_code == 200
    batch_id = batch_response.json()["batch_id"]

    response = client.post(f"/api/v1/modules/master-data/batches/{batch_id}/map", headers=admin_header)

    assert response.status_code == 200
    payload = response.json()["records"][0]["payload"]
    assert payload["CITY"] == "Fallback City"


def test_master_data_dynamic_template_build_csv_and_export_package_preserves_otm_csv_shape(
    client,
    admin_header,
):
    client.post(
        "/api/v1/modules/master-data/templates/drafts",
        json=dynamic_locations_template_payload("LOCATIONS_DYNAMIC_CSV"),
        headers=admin_header,
    )
    publish = client.post(
        "/api/v1/modules/master-data/templates/LOCATIONS_DYNAMIC_CSV/publish",
        headers=admin_header,
    )
    assert publish.status_code == 200
    workbook = Workbook()
    locations = workbook.active
    locations.title = "LOCATIONS"
    locations.append(["Location ID", "Location Name"])
    locations.append(["SYN.LOCATION_001", "Synthetic Location"])
    workbook_bytes = BytesIO()
    workbook.save(workbook_bytes)
    workbook_bytes.seek(0)
    batch_response = client.post(
        "/api/v1/modules/master-data/templates/LOCATIONS_DYNAMIC_CSV/batches",
        headers=admin_header,
        files={
            "file": (
                "locations_dynamic_csv.xlsx",
                workbook_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert batch_response.status_code == 200
    batch_id = batch_response.json()["batch_id"]
    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/map", headers=admin_header)
    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/build-output", headers=admin_header)

    csv_response = client.post(
        f"/api/v1/modules/master-data/batches/{batch_id}/build-csv",
        headers=admin_header,
    )

    assert csv_response.status_code == 200
    csv_payload = csv_response.json()
    assert csv_payload["csv_file_count"] == 1
    content_lines = csv_payload["files"][0]["content"].splitlines()
    assert content_lines[0] == "LOCATION"
    assert content_lines[1].split(",") == [
        "LOCATION_GID",
        "LOCATION_XID",
        "LOCATION_NAME",
        "COUNTRY_CODE3_GID",
    ]
    assert content_lines[2].split(",") == [
        "SYN.LOCATION_001",
        "SYN.LOCATION_001",
        "Synthetic Location",
        "USA",
    ]

    export_response = client.post(
        f"/api/v1/modules/master-data/batches/{batch_id}/export-csv-package",
        headers=admin_header,
    )

    assert export_response.status_code == 200
    export_payload = export_response.json()
    assert export_payload["file_name"].startswith("master_data_batch_")
    assert export_payload["tables"] == ["LOCATION"]


def test_master_data_dynamic_template_date_column_csv_includes_alter_session(
    client,
    admin_header,
    db_session,
):
    payload = {
        "code": "ITEMS_DATE_CSV",
        "name": "Items Date CSV",
        "catalog_macro_object_code": "ITEM",
        "data_category": "MASTER_DATA",
        "target_tables": [{"table_name": "ITEM", "sequence": 10, "required": True}],
        "sheets": [
            {
                "code": "ITEMS",
                "name": "Items",
                "sequence": 10,
                "field_keys": ["item_gid", "effective_date"],
            }
        ],
        "fields": [
            {
                "field_key": "item_gid",
                "label": "Item GID",
                "data_type": "string",
                "required": True,
                "sheet_code": "ITEMS",
            },
            {
                "field_key": "effective_date",
                "label": "Effective Date",
                "data_type": "date",
                "required": False,
                "sheet_code": "ITEMS",
            },
        ],
        "mappings": [
            {
                "mapping_key": "item_gid_to_gid",
                "source_type": "USER_FIELD",
                "source_field_key": "item_gid",
                "target_table": "ITEM",
                "target_column": "ITEM_GID",
                "required": True,
            },
            {
                "mapping_key": "item_gid_to_xid",
                "source_type": "USER_FIELD",
                "source_field_key": "item_gid",
                "target_table": "ITEM",
                "target_column": "ITEM_XID",
                "required": True,
            },
            {
                "mapping_key": "effective_date_to_item",
                "source_type": "USER_FIELD",
                "source_field_key": "effective_date",
                "target_table": "ITEM",
                "target_column": "EFFECTIVE_DATE",
                "required": False,
            },
        ],
        "relationship_rules": [],
        "documentation_refs": [
            {
                "source_type": "DATA_DICTIONARY",
                "scope": "ITEM",
                "note": "Validated against local OTM Data Dictionary.",
            }
        ],
    }
    create = client.post("/api/v1/modules/master-data/templates/drafts", json=payload, headers=admin_header)
    assert create.status_code == 200
    publish = client.post("/api/v1/modules/master-data/templates/ITEMS_DATE_CSV/publish", headers=admin_header)
    assert publish.status_code == 200

    workbook = Workbook()
    items = workbook.active
    items.title = "ITEMS"
    items.append(["Item GID", "Effective Date"])
    items.append(["OTM1.ITEM_DATE_001", "2026-05-23 09:30:00"])
    workbook_bytes = BytesIO()
    workbook.save(workbook_bytes)
    workbook_bytes.seek(0)
    batch_response = client.post(
        "/api/v1/modules/master-data/templates/ITEMS_DATE_CSV/batches",
        headers=admin_header,
        files={
            "file": (
                "items_date_csv.xlsx",
                workbook_bytes.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert batch_response.status_code == 200
    batch_id = batch_response.json()["batch_id"]

    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/map", headers=admin_header)
    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/build-output", headers=admin_header)
    csv_response = client.post(f"/api/v1/modules/master-data/batches/{batch_id}/build-csv", headers=admin_header)
    export_response = client.post(
        f"/api/v1/modules/master-data/batches/{batch_id}/export-csv-package",
        headers=admin_header,
    )

    assert csv_response.status_code == 200
    csv_payload = csv_response.json()
    assert csv_payload["csv_file_count"] == 1
    lines = csv_payload["files"][0]["content"].splitlines()
    assert lines[0] == "ITEM"
    assert lines[1] == "ITEM_GID,ITEM_XID,EFFECTIVE_DATE"
    assert lines[2] == "exec alter session set nls_date_format = 'YYYY-MM-DD HH24:MI:SS'"
    assert lines[3] == "OTM1.ITEM_DATE_001,OTM1.ITEM_DATE_001,2026-05-23 09:30:00"

    assert export_response.status_code == 200
    artifact = db_session.query(Artifact).filter(Artifact.id == export_response.json()["artifact_id"]).one()
    with zipfile.ZipFile(artifact.file_path) as archive:
        csv_text = archive.read("csv/001_ITEM.csv").decode("utf-8")
    assert csv_text.splitlines()[2] == "exec alter session set nls_date_format = 'YYYY-MM-DD HH24:MI:SS'"


def test_master_data_dynamic_template_relationship_rules_validate_orphans_from_v2_definition(
    client,
    admin_header,
):
    client.post(
        "/api/v1/modules/master-data/templates/drafts",
        json=dynamic_locations_with_addresses_payload("LOCATIONS_DYNAMIC_RELATIONSHIP"),
        headers=admin_header,
    )
    publish = client.post(
        "/api/v1/modules/master-data/templates/LOCATIONS_DYNAMIC_RELATIONSHIP/publish",
        headers=admin_header,
    )
    assert publish.status_code == 200
    workbook = Workbook()
    locations = workbook.active
    locations.title = "LOCATIONS"
    locations.append(["Location ID", "Location Name"])
    locations.append(["SYN.LOCATION_001", "Synthetic Location"])
    addresses = workbook.create_sheet("LOCATION_ADDRESSES")
    addresses.append(["Address Location ID", "Address Line"])
    addresses.append(["SYN.LOCATION_999", "Synthetic Address"])
    workbook_bytes = BytesIO()
    workbook.save(workbook_bytes)
    workbook_bytes.seek(0)
    batch_response = client.post(
        "/api/v1/modules/master-data/templates/LOCATIONS_DYNAMIC_RELATIONSHIP/batches",
        headers=admin_header,
        files={
            "file": (
                "locations_dynamic_relationship.xlsx",
                workbook_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert batch_response.status_code == 200
    batch_id = batch_response.json()["batch_id"]

    response = client.post(
        f"/api/v1/modules/master-data/batches/{batch_id}/validate-relationships",
        headers=admin_header,
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "RELATIONSHIP_FAILED"
    assert payload["issues"][0]["code"] == "MASTER_DATA_RELATIONSHIP_ORPHAN"
    assert payload["issues"][0]["child_field_name"] == "address_location_gid"
    assert payload["issues"][0]["missing_value"] == "SYN.LOCATION_999"


def test_master_data_operational_location_mapping_runs_full_backend_flow_and_blocks_early_mapping(
    client,
    admin_header,
):
    client.post(
        "/api/v1/modules/master-data/templates/drafts",
        json=dynamic_operational_locations_payload(),
        headers=admin_header,
    )
    publish = client.post(
        "/api/v1/modules/master-data/templates/LOCATIONS_OPERATIONAL_QA/publish",
        headers=admin_header,
    )
    assert publish.status_code == 200
    assert publish.json()["validation"]["valid"] is True

    workbook = Workbook()
    locations = workbook.active
    locations.title = "LOCATIONS"
    locations.append(
        [
            "Location GID",
            "Location XID",
            "Location Name",
            "City",
            "Province Code",
            "Postal Code",
            "Country Code",
            "Time Zone",
            "Latitude",
            "Longitude",
            "Equipment Restriction Profile",
            "Appointment Activity Type",
        ]
    )
    locations.append(
        [
            "SYN.LOC_DC_001",
            "LOC_DC_001",
            "Synthetic Distribution Center",
            "Atlanta",
            "GA",
            "30301",
            "USA",
            "America/New_York",
            "33.74900",
            "-84.38800",
            "SYN.EGP_DRYVAN",
            "LIVE",
        ]
    )
    addresses = workbook.create_sheet("LOCATION_ADDRESSES")
    addresses.append(["Address Location GID", "Address Line Sequence", "Address Line"])
    addresses.append(["SYN.LOC_DC_001", "1", "100 Synthetic Logistics Way"])
    capacities = workbook.create_sheet("LOCATION_CAPACITIES")
    capacities.append(["Capacity Location GID", "Location Capacity GID", "Location Capacity XID", "Capacity Calendar GID"])
    capacities.append(["SYN.LOC_DC_001", "SYN.LOC_CAP_DC_001", "LOC_CAP_DC_001", "SYN.CAL_DOCK_WEEKDAY"])
    activity = workbook.create_sheet("LOCATION_ACTIVITY_TIMES")
    activity.append(["Activity Location GID", "Location Role GID", "Activity Time Definition GID"])
    activity.append(["SYN.LOC_DC_001", "SYN.LOC_ROLE_SHIPPER", "SYN.ACT_TIME_LIVE_90"])
    docks = workbook.create_sheet("LOCATION_DOCKS")
    docks.append(["Dock Location GID", "Dock Code", "Dock Equipment Profile", "Is Load Dock", "Load Sequence", "Is Unload Dock", "Unload Sequence"])
    docks.append(["SYN.LOC_DC_001", "DOCK_A", "SYN.EGP_DRYVAN", "Y", "10", "Y", "10"])
    profiles = workbook.create_sheet("EQUIPMENT_PROFILES")
    profiles.append(["Equipment Profile GID", "Equipment Profile XID", "Equipment Profile Name"])
    profiles.append(["SYN.EGP_DRYVAN", "EGP_DRYVAN", "Synthetic Dry Van Equipment"])
    profile_details = workbook.create_sheet("EQUIPMENT_PROFILE_DETAILS")
    profile_details.append(["Detail Equipment Profile GID", "Equipment Group GID"])
    profile_details.append(["SYN.EGP_DRYVAN", "SYN.EQG_53_DRYVAN"])
    workbook_bytes = BytesIO()
    workbook.save(workbook_bytes)
    workbook_bytes.seek(0)

    batch_response = client.post(
        "/api/v1/modules/master-data/templates/LOCATIONS_OPERATIONAL_QA/batches",
        headers=admin_header,
        files={
            "file": (
                "locations_operational_qa.xlsx",
                workbook_bytes.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert batch_response.status_code == 200
    batch_id = batch_response.json()["batch_id"]

    early_map = client.post(f"/api/v1/modules/master-data/batches/{batch_id}/map", headers=admin_header)
    assert early_map.status_code == 409
    assert early_map.json()["code"] == "MASTER_DATA_BATCH_NOT_MAPPABLE"

    relationship = client.post(
        f"/api/v1/modules/master-data/batches/{batch_id}/validate-relationships",
        headers=admin_header,
    )
    assert relationship.status_code == 200
    assert relationship.json()["status"] == "RELATIONSHIP_VALIDATED"

    mapped = client.post(f"/api/v1/modules/master-data/batches/{batch_id}/map", headers=admin_header)
    assert mapped.status_code == 200
    assert mapped.json()["canonical_record_count"] == 7
    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/build-output", headers=admin_header)
    csv_response = client.post(f"/api/v1/modules/master-data/batches/{batch_id}/build-csv", headers=admin_header)

    assert csv_response.status_code == 200
    files = {csv_file["table_name"]: csv_file["content"].splitlines() for csv_file in csv_response.json()["files"]}
    assert set(files) == {
        "LOCATION",
        "LOCATION_ADDRESS",
        "LOCATION_CAPACITY",
        "LOCATION_ACTIVITY_TIME_DEF",
        "LOCATION_LOAD_UNLOAD_POINT",
        "EQUIPMENT_GROUP_PROFILE",
        "EQUIPMENT_GROUP_PROFILE_D",
    }
    assert files["LOCATION"][0] == "LOCATION"
    assert "SHUTTLE_LOT_EQ_GRP_PROFILE_GID" in files["LOCATION"][1].split(",")
    assert files["LOCATION_LOAD_UNLOAD_POINT"][2].split(",")[:3] == [
        "SYN.LOC_DC_001",
        "DOCK_A",
        "SYN.EGP_DRYVAN",
    ]


def test_master_data_item_packaging_mapping_runs_full_backend_flow(client, admin_header):
    client.post(
        "/api/v1/modules/master-data/templates/drafts",
        json=dynamic_item_packaging_payload(),
        headers=admin_header,
    )
    publish = client.post(
        "/api/v1/modules/master-data/templates/ITEM_PACKAGING_OPERATIONAL_QA/publish",
        headers=admin_header,
    )
    assert publish.status_code == 200
    assert publish.json()["validation"]["valid"] is True

    workbook = Workbook()
    items = workbook.active
    items.title = "ITEMS"
    items.append(["Item GID", "Item XID", "Item Name", "Item Type GID", "Unit of Measure"])
    items.append(["SYN.ITEM_WIDGET_001", "ITEM_WIDGET_001", "Synthetic Widget", "SYN.ITEM_TYPE_GENERAL", "EA"])
    specs = workbook.create_sheet("SHIP_UNIT_SPECS")
    specs.append(
        [
            "Ship Unit Spec GID",
            "Ship Unit Spec XID",
            "Ship Unit Spec Name",
            "Unit Type",
            "Length",
            "Length UOM",
            "Width",
            "Width UOM",
            "Height",
            "Height UOM",
            "Tare Weight",
            "Tare Weight UOM",
            "Effective Volume",
            "Effective Volume UOM",
            "Inside/On Max",
        ]
    )
    specs.append(["SYN.SUS_CASE", "SUS_CASE", "Synthetic Case", "P", "12", "IN", "10", "IN", "8", "IN", "1.5", "LB", "0.55", "CUFT", "I"])
    specs.append(["SYN.SUS_PALLET", "SUS_PALLET", "Synthetic Pallet", "T", "48", "IN", "40", "IN", "60", "IN", "45", "LB", "66.67", "CUFT", "O"])
    packaged_items = workbook.create_sheet("PACKAGED_ITEMS")
    packaged_items.append(
        [
            "Packaged Item GID",
            "Packaged Item XID",
            "Item GID",
            "Packaging Unit GID",
            "Package Weight",
            "Package Weight UOM",
            "Package Volume",
            "Package Volume UOM",
            "Package Length",
            "Package Length UOM",
            "Package Width",
            "Package Width UOM",
            "Package Height",
            "Package Height UOM",
        ]
    )
    packaged_items.append(["SYN.PKG_WIDGET_CASE", "PKG_WIDGET_CASE", "SYN.ITEM_WIDGET_001", "SYN.SUS_CASE", "24", "LB", "0.55", "CUFT", "12", "IN", "10", "IN", "8", "IN"])
    tihi = workbook.create_sheet("TI_HI")
    tihi.append(["TiHi Sequence", "Packaged Item GID", "Packaging Unit GID", "Transport Handling Unit GID", "Number of Layers", "Quantity per Layer"])
    tihi.append(["10", "SYN.PKG_WIDGET_CASE", "SYN.SUS_CASE", "SYN.SUS_PALLET", "4", "6"])
    workbook_bytes = BytesIO()
    workbook.save(workbook_bytes)
    workbook_bytes.seek(0)

    batch_response = client.post(
        "/api/v1/modules/master-data/templates/ITEM_PACKAGING_OPERATIONAL_QA/batches",
        headers=admin_header,
        files={
            "file": (
                "item_packaging_operational_qa.xlsx",
                workbook_bytes.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert batch_response.status_code == 200
    batch_id = batch_response.json()["batch_id"]
    relationship = client.post(
        f"/api/v1/modules/master-data/batches/{batch_id}/validate-relationships",
        headers=admin_header,
    )
    assert relationship.status_code == 200
    assert relationship.json()["issue_count"] == 0

    mapped = client.post(f"/api/v1/modules/master-data/batches/{batch_id}/map", headers=admin_header)
    assert mapped.status_code == 200
    assert mapped.json()["canonical_record_count"] == 5
    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/build-output", headers=admin_header)
    csv_response = client.post(f"/api/v1/modules/master-data/batches/{batch_id}/build-csv", headers=admin_header)

    assert csv_response.status_code == 200
    files = {csv_file["table_name"]: csv_file["content"].splitlines() for csv_file in csv_response.json()["files"]}
    assert set(files) == {"ITEM", "SHIP_UNIT_SPEC", "PACKAGED_ITEM", "TI_HI"}
    assert "TRANSPORT_HANDLING_UNIT_GID" in files["TI_HI"][1].split(",")
    assert files["TI_HI"][2].split(",") == ["10", "SYN.PKG_WIDGET_CASE", "SYN.SUS_CASE", "SYN.SUS_PALLET", "4", "6"]


def test_master_data_items_packaging_template_detail_and_validation(client, admin_header):
    detail = client.get(
        "/api/v1/modules/master-data/templates/ITEMS_PACKAGING_STANDARD",
        headers=admin_header,
    )

    assert detail.status_code == 200
    payload = detail.json()
    assert payload["code"] == "ITEMS_PACKAGING_STANDARD"
    assert [sheet["code"] for sheet in payload["sheets"]] == ["ITEMS", "PACKAGING", "TI_HI"]
    assert payload["sheets"][0]["target_table"] == "ITEM"
    assert [field["target_column"] for field in payload["sheets"][0]["fields"]] == [
        "ITEM_GID",
        "ITEM_XID",
        "DESCRIPTION",
        "ITEM_TYPE_GID",
    ]
    assert payload["sheets"][1]["target_table"] == "PACKAGED_ITEM"
    assert payload["sheets"][2]["target_table"] == "TI_HI"

    validation = client.post(
        "/api/v1/modules/master-data/templates/ITEMS_PACKAGING_STANDARD/validate",
        headers=admin_header,
    )

    assert validation.status_code == 200
    validation_payload = validation.json()
    assert validation_payload["valid"] is True
    assert validation_payload["issues"] == []
    assert validation_payload["summary"] == {
        "sheet_count": 3,
        "field_count": 18,
        "validated_table_count": 3,
        "validated_column_count": 18,
    }


def test_master_data_workbook_editor_contract_uses_backend_template_definition(client, admin_header):
    response = client.get(
        "/api/v1/modules/master-data/templates/ITEMS_PACKAGING_STANDARD/workbook-editor",
        headers=admin_header,
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["template_code"] == "ITEMS_PACKAGING_STANDARD"
    assert [sheet["code"] for sheet in payload["sheets"]] == ["ITEMS", "PACKAGING", "TI_HI"]
    assert payload["sheets"][0]["target_table"] == "ITEM"
    assert payload["sheets"][0]["fields"][0] == {
        "field_key": "item_gid",
        "label": "Item GID",
        "data_type": "string",
        "required": True,
    }
    assert payload["sheets"][0]["starter_rows"] == [
        {"row_id": "ITEMS-1", "values": {"item_gid": "", "item_xid": "", "description": "", "item_type_gid": ""}}
    ]
    assert payload["relationship_rules"][0]["parent_sheet_code"] == "ITEMS"


def test_master_data_workbook_editor_validation_reports_required_fields(client, admin_header):
    response = client.post(
        "/api/v1/modules/master-data/templates/ITEMS_PACKAGING_STANDARD/workbook-editor/validate",
        headers=admin_header,
        json={
            "file_name": "items_editor.xlsx",
            "sheets": [{"sheet_code": "ITEMS", "rows": [{"row_id": "ITEMS-1", "values": {"item_gid": ""}}]}],
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["valid"] is False
    assert payload["status"] == "INVALID"
    assert payload["issues"][0]["code"] == "REQUIRED_FIELD_MISSING"
    assert payload["issues"][0]["field_key"] == "item_gid"
    assert payload["issues"][0]["row_id"] == "ITEMS-1"


def test_master_data_workbook_editor_validation_reports_relationship_orphans(client, admin_header):
    response = client.post(
        "/api/v1/modules/master-data/templates/ITEMS_PACKAGING_STANDARD/workbook-editor/validate",
        headers=admin_header,
        json={
            "file_name": "items_editor.xlsx",
            "sheets": [
                {
                    "sheet_code": "ITEMS",
                    "rows": [
                        {
                            "row_id": "ITEMS-1",
                            "values": {"item_gid": "SYN.ITEM_1", "item_xid": "ITEM_1"},
                        }
                    ],
                },
                {
                    "sheet_code": "PACKAGING",
                    "rows": [
                        {
                            "row_id": "PACKAGING-1",
                            "values": {
                                "packaged_item_gid": "SYN.PKG_1",
                                "packaged_item_xid": "PKG_1",
                                "item_gid": "SYN.MISSING_ITEM",
                            },
                        }
                    ],
                },
            ],
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["valid"] is False
    assert any(
        issue["code"] == "RELATIONSHIP_PARENT_NOT_FOUND"
        and issue["sheet_code"] == "PACKAGING"
        and issue["field_key"] == "item_gid"
        for issue in payload["issues"]
    )


def test_master_data_locations_template_detail_and_validation(client, admin_header):
    detail = client.get(
        "/api/v1/modules/master-data/templates/LOCATIONS_BASIC",
        headers=admin_header,
    )

    assert detail.status_code == 200
    payload = detail.json()
    assert payload["code"] == "LOCATIONS_BASIC"
    assert [sheet["code"] for sheet in payload["sheets"]] == ["LOCATIONS", "LOCATION_ADDRESSES"]
    assert payload["sheets"][0]["target_table"] == "LOCATION"
    assert [field["target_column"] for field in payload["sheets"][0]["fields"]] == [
        "LOCATION_GID",
        "LOCATION_XID",
        "LOCATION_NAME",
        "CITY",
        "PROVINCE_CODE",
        "POSTAL_CODE",
        "COUNTRY_CODE3_GID",
        "LAT",
        "LON",
    ]
    assert payload["sheets"][1]["target_table"] == "LOCATION_ADDRESS"
    assert [field["target_column"] for field in payload["sheets"][1]["fields"]] == [
        "LOCATION_GID",
        "LINE_SEQUENCE",
        "ADDRESS_LINE",
    ]

    validation = client.post(
        "/api/v1/modules/master-data/templates/LOCATIONS_BASIC/validate",
        headers=admin_header,
    )

    assert validation.status_code == 200
    validation_payload = validation.json()
    assert validation_payload["valid"] is True
    assert validation_payload["issues"] == []
    assert validation_payload["summary"] == {
        "sheet_count": 2,
        "field_count": 12,
        "validated_table_count": 2,
        "validated_column_count": 12,
    }


def test_master_data_template_build_workbook_creates_artifact(client, admin_header, db_session):
    response = client.post(
        "/api/v1/modules/master-data/templates/REGIONS_BASIC/build-workbook",
        headers=admin_header,
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["template_code"] == "REGIONS_BASIC"
    assert payload["file_name"] == "regions_basic_v1.xlsx"
    assert payload["content_type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert payload["sheet_count"] == 2
    assert payload["field_count"] == 5

    artifact = db_session.query(Artifact).filter(Artifact.id == payload["artifact_id"]).one()
    assert artifact.source_module == "master_data"
    assert artifact.artifact_type == "master_data_template_workbook"
    assert artifact.sensitivity_level == "client_safe"

    workbook_path = Path(artifact.file_path)
    assert workbook_path.exists()
    workbook = load_workbook(workbook_path)
    assert workbook.sheetnames == ["REGIONS", "REGION_DETAILS"]
    assert [cell.value for cell in workbook["REGIONS"][1]] == ["Region GID", "Region XID", "Region Name"]
    assert [cell.value for cell in workbook["REGION_DETAILS"][1]] == ["Region GID", "Location GID"]


def test_master_data_template_batch_upload_parses_workbook(client, admin_header, db_session):
    workbook = Workbook()
    regions = workbook.active
    regions.title = "REGIONS"
    regions.append(["Region GID", "Region XID", "Region Name"])
    regions.append(["SYN.REGION_001", "REGION_001", "Synthetic Region"])
    details = workbook.create_sheet("REGION_DETAILS")
    details.append(["Region GID", "Location GID"])
    details.append(["SYN.REGION_001", "SYN.LOCATION_001"])
    workbook_bytes = BytesIO()
    workbook.save(workbook_bytes)
    workbook_bytes.seek(0)

    response = client.post(
        "/api/v1/modules/master-data/templates/REGIONS_BASIC/batches",
        headers=admin_header,
        files={
            "file": (
                "regions_basic_upload.xlsx",
                workbook_bytes.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["template_code"] == "REGIONS_BASIC"
    assert payload["status"] == "PARSED"
    assert payload["file_name"] == "regions_basic_upload.xlsx"
    assert payload["sheet_count"] == 2
    assert payload["row_count"] == 2
    assert payload["issue_count"] == 0
    assert payload["sheet_summaries"] == [
        {"sheet_code": "REGIONS", "target_table": "REGION", "row_count": 1},
        {"sheet_code": "REGION_DETAILS", "target_table": "REGION_DETAIL", "row_count": 1},
    ]

    row = db_session.execute(
        text(
            "select template_code, status, file_name, row_count, issue_count "
            "from master_data_batches where id = :batch_id"
        ),
        {"batch_id": payload["batch_id"]},
    ).one()
    assert row.template_code == "REGIONS_BASIC"
    assert row.status == "PARSED"
    assert row.file_name == "regions_basic_upload.xlsx"
    assert row.row_count == 2
    assert row.issue_count == 0


def test_master_data_batch_detail_exposes_backend_owned_available_actions(client, admin_header):
    workbook = Workbook()
    regions = workbook.active
    regions.title = "REGIONS"
    regions.append(["Region GID", "Region XID", "Region Name"])
    regions.append(["SYN.REGION_ACTION", "REGION_ACTION", "Synthetic Action Region"])
    details = workbook.create_sheet("REGION_DETAILS")
    details.append(["Region GID", "Location GID"])
    details.append(["SYN.REGION_ACTION", "SYN.LOCATION_ACTION"])
    workbook_bytes = BytesIO()
    workbook.save(workbook_bytes)
    workbook_bytes.seek(0)
    batch_response = client.post(
        "/api/v1/modules/master-data/templates/REGIONS_BASIC/batches",
        headers=admin_header,
        files={
            "file": (
                "regions_basic_actions_upload.xlsx",
                workbook_bytes.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    batch_id = batch_response.json()["batch_id"]

    parsed_actions = client.get(f"/api/v1/modules/master-data/batches/{batch_id}", headers=admin_header).json()[
        "available_actions"
    ]
    assert action_by_key(parsed_actions, "validate_relationships")["disabled"] is False
    assert action_by_key(parsed_actions, "validate_relationships")["recommended"] is True
    assert action_by_key(parsed_actions, "map_records")["disabled"] is True
    assert action_by_key(parsed_actions, "map_records")["disabled_reason"] == "RELATIONSHIP_VALIDATION_REQUIRED"
    assert action_by_key(parsed_actions, "register_load_plan_package")["disabled_reason"] == "EXPORT_REQUIRED"

    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/validate-relationships", headers=admin_header)
    validated_actions = client.get(f"/api/v1/modules/master-data/batches/{batch_id}", headers=admin_header).json()[
        "available_actions"
    ]
    assert action_by_key(validated_actions, "validate_relationships")["disabled"] is True
    assert action_by_key(validated_actions, "map_records")["disabled"] is False
    assert action_by_key(validated_actions, "map_records")["recommended"] is True

    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/map", headers=admin_header)
    mapped_actions = client.get(f"/api/v1/modules/master-data/batches/{batch_id}", headers=admin_header).json()[
        "available_actions"
    ]
    assert action_by_key(mapped_actions, "build_output")["disabled"] is False
    assert action_by_key(mapped_actions, "build_output")["recommended"] is True
    assert action_by_key(mapped_actions, "build_csv")["disabled"] is True

    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/build-output", headers=admin_header)
    output_actions = client.get(f"/api/v1/modules/master-data/batches/{batch_id}", headers=admin_header).json()[
        "available_actions"
    ]
    assert action_by_key(output_actions, "build_csv")["disabled"] is False
    assert action_by_key(output_actions, "build_csv")["recommended"] is True
    assert action_by_key(output_actions, "export_csv_package")["disabled"] is True

    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/build-csv", headers=admin_header)
    csv_actions = client.get(f"/api/v1/modules/master-data/batches/{batch_id}", headers=admin_header).json()[
        "available_actions"
    ]
    assert action_by_key(csv_actions, "export_csv_package")["disabled"] is False
    assert action_by_key(csv_actions, "export_csv_package")["recommended"] is True

    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/export-csv-package", headers=admin_header)
    exported_actions = client.get(f"/api/v1/modules/master-data/batches/{batch_id}", headers=admin_header).json()[
        "available_actions"
    ]
    assert action_by_key(exported_actions, "register_load_plan_package")["disabled"] is False
    assert action_by_key(exported_actions, "register_load_plan_package")["recommended"] is True


def test_master_data_available_actions_allow_mapping_parsed_batch_without_relationship_rules(client, admin_header):
    client.post(
        "/api/v1/modules/master-data/templates/drafts",
        json=dynamic_locations_template_payload("LOCATIONS_ACTIONS_NO_REL"),
        headers=admin_header,
    )
    client.post("/api/v1/modules/master-data/templates/LOCATIONS_ACTIONS_NO_REL/publish", headers=admin_header)

    workbook = Workbook()
    locations = workbook.active
    locations.title = "LOCATIONS"
    locations.append(["Location ID", "Location Name"])
    locations.append(["SYN.LOCATION_ACTION", "Synthetic Action Location"])
    workbook_bytes = BytesIO()
    workbook.save(workbook_bytes)
    workbook_bytes.seek(0)
    batch_response = client.post(
        "/api/v1/modules/master-data/templates/LOCATIONS_ACTIONS_NO_REL/batches",
        headers=admin_header,
        files={
            "file": (
                "locations_actions_no_rel.xlsx",
                workbook_bytes.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert batch_response.status_code == 200
    actions = batch_response.json()["available_actions"]
    assert action_by_key(actions, "validate_relationships")["disabled"] is False
    assert action_by_key(actions, "map_records")["disabled"] is False
    assert action_by_key(actions, "map_records")["disabled_reason"] is None
    assert action_by_key(actions, "map_records")["recommended"] is True


def test_master_data_template_batch_upload_skips_generated_metadata_rows(
    client,
    admin_header,
):
    workbook = Workbook()
    regions = workbook.active
    regions.title = "REGIONS"
    regions.append(["Region GID", "Region XID", "Region Name"])
    regions.append(["region_gid", "region_xid", "region_name"])
    regions.append(["REGION_GID", "REGION_XID", "REGION_NAME"])
    regions.append(["SYN.REGION_001", "REGION_001", "Synthetic Region"])
    details = workbook.create_sheet("REGION_DETAILS")
    details.append(["Region GID", "Location GID"])
    details.append(["region_gid", "location_gid"])
    details.append(["REGION_GID", "LOCATION_GID"])
    details.append(["SYN.REGION_001", "SYN.LOCATION_001"])
    workbook_bytes = BytesIO()
    workbook.save(workbook_bytes)
    workbook_bytes.seek(0)

    response = client.post(
        "/api/v1/modules/master-data/templates/REGIONS_BASIC/batches",
        headers=admin_header,
        files={
            "file": (
                "regions_basic_generated_upload.xlsx",
                workbook_bytes.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "PARSED"
    assert payload["row_count"] == 2
    assert payload["sheet_summaries"] == [
        {"sheet_code": "REGIONS", "target_table": "REGION", "row_count": 1},
        {"sheet_code": "REGION_DETAILS", "target_table": "REGION_DETAIL", "row_count": 1},
    ]


def test_master_data_template_batch_upload_persists_parse_issues(client, admin_header, db_session):
    workbook = Workbook()
    regions = workbook.active
    regions.title = "REGIONS"
    regions.append(["Wrong Region", "Region XID", "Region Name"])
    regions.append(["SYN.REGION_001", "REGION_001", "Synthetic Region"])
    details = workbook.create_sheet("REGION_DETAILS")
    details.append(["Region GID", "Location GID"])
    details.append(["SYN.REGION_001", "SYN.LOCATION_001"])
    workbook_bytes = BytesIO()
    workbook.save(workbook_bytes)
    workbook_bytes.seek(0)

    response = client.post(
        "/api/v1/modules/master-data/templates/REGIONS_BASIC/batches",
        headers=admin_header,
        files={
            "file": (
                "regions_basic_bad_headers.xlsx",
                workbook_bytes.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["template_code"] == "REGIONS_BASIC"
    assert payload["status"] == "PARSE_FAILED"
    assert payload["file_name"] == "regions_basic_bad_headers.xlsx"
    assert payload["row_count"] == 0
    assert payload["issue_count"] == 1
    assert payload["issues"] == [
        {
            "code": "MASTER_DATA_WORKBOOK_HEADERS_INVALID",
            "severity": "ERROR",
            "sheet_code": "REGIONS",
            "message": "Uploaded workbook headers do not match the template.",
            "expected_headers": ["Region GID", "Region XID", "Region Name"],
            "actual_headers": ["Wrong Region", "Region XID", "Region Name"],
        }
    ]

    row = db_session.execute(
        text(
            "select status, row_count, issue_count, issues_json "
            "from master_data_batches where id = :batch_id"
        ),
        {"batch_id": payload["batch_id"]},
    ).one()
    assert row.status == "PARSE_FAILED"
    assert row.row_count == 0
    assert row.issue_count == 1
    assert "MASTER_DATA_WORKBOOK_HEADERS_INVALID" in row.issues_json


def test_master_data_batch_relationship_validation_persists_orphan_issues(
    client,
    admin_header,
    db_session,
):
    workbook = Workbook()
    regions = workbook.active
    regions.title = "REGIONS"
    regions.append(["Region GID", "Region XID", "Region Name"])
    regions.append(["SYN.REGION_001", "REGION_001", "Synthetic Region"])
    details = workbook.create_sheet("REGION_DETAILS")
    details.append(["Region GID", "Location GID"])
    details.append(["SYN.REGION_999", "SYN.LOCATION_001"])
    workbook_bytes = BytesIO()
    workbook.save(workbook_bytes)
    workbook_bytes.seek(0)
    batch_response = client.post(
        "/api/v1/modules/master-data/templates/REGIONS_BASIC/batches",
        headers=admin_header,
        files={
            "file": (
                "regions_basic_orphan_detail.xlsx",
                workbook_bytes.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    batch_id = batch_response.json()["batch_id"]

    response = client.post(
        f"/api/v1/modules/master-data/batches/{batch_id}/validate-relationships",
        headers=admin_header,
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["batch_id"] == batch_id
    assert payload["status"] == "RELATIONSHIP_FAILED"
    assert payload["issue_count"] == 1
    assert payload["issues"] == [
        {
            "code": "MASTER_DATA_RELATIONSHIP_ORPHAN",
            "severity": "ERROR",
            "sheet_code": "REGION_DETAILS",
            "field_name": "region_gid",
            "row_number": 2,
            "message": "Child row references a parent key that does not exist in the same batch.",
            "parent_sheet_code": "REGIONS",
            "parent_field_name": "region_gid",
            "missing_value": "SYN.REGION_999",
        }
    ]

    row = db_session.execute(
        text(
            "select status, issue_count, issues_json "
            "from master_data_batches where id = :batch_id"
        ),
        {"batch_id": batch_id},
    ).one()
    assert row.status == "RELATIONSHIP_FAILED"
    assert row.issue_count == 1
    assert "MASTER_DATA_RELATIONSHIP_ORPHAN" in row.issues_json


def test_master_data_batch_relationship_validation_reports_actual_workbook_row(
    client,
    admin_header,
):
    workbook = Workbook()
    regions = workbook.active
    regions.title = "REGIONS"
    regions.append(["Region GID", "Region XID", "Region Name"])
    regions.append(["region_gid", "region_xid", "region_name"])
    regions.append(["REGION_GID", "REGION_XID", "REGION_NAME"])
    regions.append(["SYN.REGION_001", "REGION_001", "Synthetic Region"])
    details = workbook.create_sheet("REGION_DETAILS")
    details.append(["Region GID", "Location GID"])
    details.append(["region_gid", "location_gid"])
    details.append(["REGION_GID", "LOCATION_GID"])
    details.append(["SYN.REGION_999", "SYN.LOCATION_001"])
    workbook_bytes = BytesIO()
    workbook.save(workbook_bytes)
    workbook_bytes.seek(0)
    batch_response = client.post(
        "/api/v1/modules/master-data/templates/REGIONS_BASIC/batches",
        headers=admin_header,
        files={
            "file": (
                "regions_basic_generated_orphan_detail.xlsx",
                workbook_bytes.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    batch_id = batch_response.json()["batch_id"]

    response = client.post(
        f"/api/v1/modules/master-data/batches/{batch_id}/validate-relationships",
        headers=admin_header,
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "RELATIONSHIP_FAILED"
    assert payload["issues"][0]["row_number"] == 4


def test_master_data_batch_relationship_validation_is_idempotent(client, admin_header):
    workbook = Workbook()
    regions = workbook.active
    regions.title = "REGIONS"
    regions.append(["Region GID", "Region XID", "Region Name"])
    regions.append(["SYN.REGION_001", "REGION_001", "Synthetic Region"])
    details = workbook.create_sheet("REGION_DETAILS")
    details.append(["Region GID", "Location GID"])
    details.append(["SYN.REGION_001", "SYN.LOCATION_001"])
    workbook_bytes = BytesIO()
    workbook.save(workbook_bytes)
    workbook_bytes.seek(0)
    batch_response = client.post(
        "/api/v1/modules/master-data/templates/REGIONS_BASIC/batches",
        headers=admin_header,
        files={
            "file": (
                "regions_basic_upload.xlsx",
                workbook_bytes.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    batch_id = batch_response.json()["batch_id"]
    first_response = client.post(
        f"/api/v1/modules/master-data/batches/{batch_id}/validate-relationships",
        headers=admin_header,
    )

    second_response = client.post(
        f"/api/v1/modules/master-data/batches/{batch_id}/validate-relationships",
        headers=admin_header,
    )

    assert first_response.status_code == 200
    assert second_response.status_code == 200
    assert second_response.json() == {
        "batch_id": batch_id,
        "status": "RELATIONSHIP_VALIDATED",
        "issue_count": 0,
        "issues": [],
    }


def test_master_data_batch_mapping_requires_relationship_validation(client, admin_header):
    workbook = Workbook()
    regions = workbook.active
    regions.title = "REGIONS"
    regions.append(["Region GID", "Region XID", "Region Name"])
    regions.append(["SYN.REGION_001", "REGION_001", "Synthetic Region"])
    details = workbook.create_sheet("REGION_DETAILS")
    details.append(["Region GID", "Location GID"])
    details.append(["SYN.REGION_001", "SYN.LOCATION_001"])
    workbook_bytes = BytesIO()
    workbook.save(workbook_bytes)
    workbook_bytes.seek(0)
    batch_response = client.post(
        "/api/v1/modules/master-data/templates/REGIONS_BASIC/batches",
        headers=admin_header,
        files={
            "file": (
                "regions_basic_upload.xlsx",
                workbook_bytes.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    batch_id = batch_response.json()["batch_id"]

    response = client.post(f"/api/v1/modules/master-data/batches/{batch_id}/map", headers=admin_header)

    assert response.status_code == 409
    payload = response.json()
    assert payload["code"] == "MASTER_DATA_BATCH_NOT_MAPPABLE"
    assert "relationship-validated" in payload["details"]["error"]


def test_master_data_batch_mapping_creates_canonical_records(client, admin_header, db_session):
    workbook = Workbook()
    regions = workbook.active
    regions.title = "REGIONS"
    regions.append(["Region GID", "Region XID", "Region Name"])
    regions.append(["SYN.REGION_001", "REGION_001", "Synthetic Region"])
    details = workbook.create_sheet("REGION_DETAILS")
    details.append(["Region GID", "Location GID"])
    details.append(["SYN.REGION_001", "SYN.LOCATION_001"])
    workbook_bytes = BytesIO()
    workbook.save(workbook_bytes)
    workbook_bytes.seek(0)
    batch_response = client.post(
        "/api/v1/modules/master-data/templates/REGIONS_BASIC/batches",
        headers=admin_header,
        files={
            "file": (
                "regions_basic_upload.xlsx",
                workbook_bytes.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    batch_id = batch_response.json()["batch_id"]
    client.post(
        f"/api/v1/modules/master-data/batches/{batch_id}/validate-relationships",
        headers=admin_header,
    )

    response = client.post(f"/api/v1/modules/master-data/batches/{batch_id}/map", headers=admin_header)

    assert response.status_code == 200
    payload = response.json()
    assert payload["batch_id"] == batch_id
    assert payload["status"] == "MAPPED"
    assert payload["canonical_record_count"] == 2
    assert payload["records"] == [
        {
            "sheet_code": "REGIONS",
            "target_table": "REGION",
            "record_index": 1,
            "payload": {
                "REGION_GID": "SYN.REGION_001",
                "REGION_XID": "REGION_001",
                "REGION_NAME": "Synthetic Region",
            },
        },
        {
            "sheet_code": "REGION_DETAILS",
            "target_table": "REGION_DETAIL",
            "record_index": 1,
            "payload": {
                "REGION_GID": "SYN.REGION_001",
                "LOCATION_GID": "SYN.LOCATION_001",
            },
        },
    ]

    rows = db_session.execute(
        text(
            "select sheet_code, target_table, record_index, payload_json "
            "from master_data_canonical_records where batch_id = :batch_id"
        ),
        {"batch_id": batch_id},
    ).all()
    assert len(rows) == 2
    rows_by_sheet = {row.sheet_code: row for row in rows}
    assert rows_by_sheet["REGIONS"].target_table == "REGION"
    assert '"REGION_GID": "SYN.REGION_001"' in rows_by_sheet["REGIONS"].payload_json


def test_master_data_batch_build_output_creates_output_records(client, admin_header, db_session):
    workbook = Workbook()
    regions = workbook.active
    regions.title = "REGIONS"
    regions.append(["Region GID", "Region XID", "Region Name"])
    regions.append(["SYN.REGION_001", "REGION_001", "Synthetic Region"])
    details = workbook.create_sheet("REGION_DETAILS")
    details.append(["Region GID", "Location GID"])
    details.append(["SYN.REGION_001", "SYN.LOCATION_001"])
    workbook_bytes = BytesIO()
    workbook.save(workbook_bytes)
    workbook_bytes.seek(0)
    batch_response = client.post(
        "/api/v1/modules/master-data/templates/REGIONS_BASIC/batches",
        headers=admin_header,
        files={
            "file": (
                "regions_basic_upload.xlsx",
                workbook_bytes.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    batch_id = batch_response.json()["batch_id"]
    client.post(
        f"/api/v1/modules/master-data/batches/{batch_id}/validate-relationships",
        headers=admin_header,
    )
    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/map", headers=admin_header)

    response = client.post(
        f"/api/v1/modules/master-data/batches/{batch_id}/build-output",
        headers=admin_header,
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["batch_id"] == batch_id
    assert payload["status"] == "OUTPUT_BUILT"
    assert payload["output_record_count"] == 2
    assert payload["records"] == [
        {
            "target_table": "REGION",
            "record_index": 1,
            "payload": {
                "REGION_GID": "SYN.REGION_001",
                "REGION_NAME": "Synthetic Region",
                "REGION_XID": "REGION_001",
            },
        },
        {
            "target_table": "REGION_DETAIL",
            "record_index": 1,
            "payload": {
                "LOCATION_GID": "SYN.LOCATION_001",
                "REGION_GID": "SYN.REGION_001",
            },
        },
    ]

    rows = db_session.execute(
        text(
            "select target_table, record_index, payload_json "
            "from master_data_output_records where batch_id = :batch_id"
        ),
        {"batch_id": batch_id},
    ).all()
    assert len(rows) == 2
    rows_by_table = {row.target_table: row for row in rows}
    assert rows_by_table["REGION"].record_index == 1
    assert '"REGION_XID": "REGION_001"' in rows_by_table["REGION"].payload_json


def test_master_data_batch_output_records_endpoint_returns_backend_preview(client, admin_header):
    workbook = Workbook()
    regions = workbook.active
    regions.title = "REGIONS"
    regions.append(["Region GID", "Region XID", "Region Name"])
    regions.append(["SYN.REGION_001", "REGION_001", "Synthetic Region"])
    details = workbook.create_sheet("REGION_DETAILS")
    details.append(["Region GID", "Location GID"])
    details.append(["SYN.REGION_001", "SYN.LOCATION_001"])
    workbook_bytes = BytesIO()
    workbook.save(workbook_bytes)
    workbook_bytes.seek(0)
    batch_response = client.post(
        "/api/v1/modules/master-data/templates/REGIONS_BASIC/batches",
        headers=admin_header,
        files={
            "file": (
                "regions_basic_upload.xlsx",
                workbook_bytes.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    batch_id = batch_response.json()["batch_id"]
    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/validate-relationships", headers=admin_header)
    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/map", headers=admin_header)
    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/build-output", headers=admin_header)

    response = client.get(f"/api/v1/modules/master-data/batches/{batch_id}/output-records", headers=admin_header)

    assert response.status_code == 200
    payload = response.json()
    assert payload["total"] == 2
    assert payload["items"][0]["target_table"] == "REGION"
    assert payload["items"][0]["payload"]["REGION_GID"] == "SYN.REGION_001"


def test_master_data_batch_build_csv_creates_otm_csv_files(client, admin_header, db_session):
    workbook = Workbook()
    regions = workbook.active
    regions.title = "REGIONS"
    regions.append(["Region GID", "Region XID", "Region Name"])
    regions.append(["SYN.REGION_001", "REGION_001", "Synthetic Region"])
    details = workbook.create_sheet("REGION_DETAILS")
    details.append(["Region GID", "Location GID"])
    details.append(["SYN.REGION_001", "SYN.LOCATION_001"])
    workbook_bytes = BytesIO()
    workbook.save(workbook_bytes)
    workbook_bytes.seek(0)
    batch_response = client.post(
        "/api/v1/modules/master-data/templates/REGIONS_BASIC/batches",
        headers=admin_header,
        files={
            "file": (
                "regions_basic_upload.xlsx",
                workbook_bytes.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    batch_id = batch_response.json()["batch_id"]
    client.post(
        f"/api/v1/modules/master-data/batches/{batch_id}/validate-relationships",
        headers=admin_header,
    )
    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/map", headers=admin_header)
    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/build-output", headers=admin_header)

    response = client.post(
        f"/api/v1/modules/master-data/batches/{batch_id}/build-csv",
        headers=admin_header,
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["batch_id"] == batch_id
    assert payload["status"] == "CSV_BUILT"
    assert payload["csv_file_count"] == 2
    assert payload["files"][0]["table_name"] == "REGION"
    assert payload["files"][0]["file_name"] == "001_REGION.csv"
    assert payload["files"][0]["row_count"] == 1
    assert payload["files"][0]["content"].splitlines() == [
        "REGION",
        "REGION_GID,REGION_XID,REGION_NAME",
        "SYN.REGION_001,REGION_001,Synthetic Region",
    ]
    assert payload["files"][1]["table_name"] == "REGION_DETAIL"
    assert payload["files"][1]["file_name"] == "002_REGION_DETAIL.csv"
    assert payload["files"][1]["content"].splitlines() == [
        "REGION_DETAIL",
        "REGION_GID,LOCATION_GID",
        "SYN.REGION_001,SYN.LOCATION_001",
    ]

    rows = db_session.execute(
        text(
            "select table_name, file_name, row_count, content "
            "from master_data_csv_files where batch_id = :batch_id order by file_name"
        ),
        {"batch_id": batch_id},
    ).all()
    assert len(rows) == 2
    assert rows[0].table_name == "REGION"
    assert rows[0].file_name == "001_REGION.csv"
    assert rows[0].row_count == 1
    assert rows[0].content.startswith("REGION\nREGION_GID,REGION_XID,REGION_NAME")


def test_master_data_batch_csv_files_endpoint_returns_preview_only(client, admin_header):
    workbook = Workbook()
    regions = workbook.active
    regions.title = "REGIONS"
    regions.append(["Region GID", "Region XID", "Region Name"])
    regions.append(["SYN.REGION_001", "REGION_001", "Synthetic Region"])
    details = workbook.create_sheet("REGION_DETAILS")
    details.append(["Region GID", "Location GID"])
    details.append(["SYN.REGION_001", "SYN.LOCATION_001"])
    workbook_bytes = BytesIO()
    workbook.save(workbook_bytes)
    workbook_bytes.seek(0)
    batch_response = client.post(
        "/api/v1/modules/master-data/templates/REGIONS_BASIC/batches",
        headers=admin_header,
        files={
            "file": (
                "regions_basic_upload.xlsx",
                workbook_bytes.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    batch_id = batch_response.json()["batch_id"]
    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/validate-relationships", headers=admin_header)
    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/map", headers=admin_header)
    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/build-output", headers=admin_header)
    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/build-csv", headers=admin_header)

    response = client.get(f"/api/v1/modules/master-data/batches/{batch_id}/csv-files", headers=admin_header)

    assert response.status_code == 200
    payload = response.json()
    assert payload["total"] == 2
    assert payload["items"][0]["file_name"] == "001_REGION.csv"
    assert payload["items"][0]["content_preview"].startswith("REGION\nREGION_GID,REGION_XID,REGION_NAME")
    assert "content" not in payload["items"][0]


def test_master_data_batch_list_filters_and_paginates(client, admin_header):
    for index in range(2):
        workbook = Workbook()
        regions = workbook.active
        regions.title = "REGIONS"
        regions.append(["Region GID", "Region XID", "Region Name"])
        regions.append([f"SYN.REGION_FILTER_{index}", f"REGION_FILTER_{index}", "Synthetic Filter Region"])
        if index == 1:
            regions.append(["SYN.REGION_FILTER_EXTRA", "REGION_FILTER_EXTRA", "Synthetic Extra Region"])
        details = workbook.create_sheet("REGION_DETAILS")
        details.append(["Region GID", "Location GID"])
        details.append([f"SYN.REGION_FILTER_{index}", f"SYN.LOCATION_FILTER_{index}"])
        if index == 1:
            details.append(["SYN.REGION_FILTER_EXTRA", "SYN.LOCATION_FILTER_EXTRA"])
        workbook_bytes = BytesIO()
        workbook.save(workbook_bytes)
        workbook_bytes.seek(0)
        response = client.post(
            "/api/v1/modules/master-data/templates/REGIONS_BASIC/batches",
            headers=admin_header,
            files={
                "file": (
                    f"regions_basic_upload_{index}.xlsx",
                    workbook_bytes.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200

    response = client.get(
        "/api/v1/modules/master-data/batches",
        params={"template_code": "REGIONS_BASIC", "status": "PARSED", "page": 1, "page_size": 1},
        headers=admin_header,
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["total"] == 2
    assert payload["page"] == 1
    assert payload["page_size"] == 1
    assert len(payload["items"]) == 1
    assert payload["items"][0]["template_code"] == "REGIONS_BASIC"
    assert payload["items"][0]["status"] == "PARSED"

    advanced_response = client.get(
        "/api/v1/modules/master-data/batches",
        params={"file_name_contains": "upload_1", "min_row_count": 4},
        headers=admin_header,
    )

    assert advanced_response.status_code == 200
    advanced_payload = advanced_response.json()
    assert advanced_payload["total"] == 1
    assert advanced_payload["items"][0]["file_name"] == "regions_basic_upload_1.xlsx"
    assert advanced_payload["items"][0]["row_count"] == 4


def test_master_data_batch_summary_groups_filtered_history(client, admin_header):
    for index in range(2):
        workbook = Workbook()
        regions = workbook.active
        regions.title = "REGIONS"
        regions.append(["Region GID", "Region XID", "Region Name"])
        regions.append([f"SYN.REGION_SUMMARY_{index}", f"REGION_SUMMARY_{index}", "Synthetic Summary Region"])
        if index == 1:
            regions.append(["SYN.REGION_SUMMARY_EXTRA", "REGION_SUMMARY_EXTRA", "Synthetic Extra Region"])
        details = workbook.create_sheet("REGION_DETAILS")
        details.append(["Region GID", "Location GID"])
        details.append([f"SYN.REGION_SUMMARY_{index}", f"SYN.LOCATION_SUMMARY_{index}"])
        if index == 1:
            details.append(["SYN.REGION_SUMMARY_EXTRA", "SYN.LOCATION_SUMMARY_EXTRA"])
        workbook_bytes = BytesIO()
        workbook.save(workbook_bytes)
        workbook_bytes.seek(0)
        response = client.post(
            "/api/v1/modules/master-data/templates/REGIONS_BASIC/batches",
            files={
                "file": (
                    f"regions_summary_upload_{index}.xlsx",
                    workbook_bytes.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers=admin_header,
        )
        assert response.status_code == 200

    response = client.get(
        "/api/v1/modules/master-data/batches/summary",
        params={"file_name_contains": "summary_upload", "min_row_count": 2},
        headers=admin_header,
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["total_batches"] == 2
    assert payload["total_rows"] == 6
    assert payload["total_issues"] == 0
    assert payload["status_breakdown"] == [{"status": "PARSED", "batch_count": 2, "row_count": 6, "issue_count": 0}]
    assert payload["template_breakdown"] == [
        {"template_code": "REGIONS_BASIC", "batch_count": 2, "row_count": 6, "issue_count": 0}
    ]


def test_master_data_batch_summary_returns_empty_totals_for_no_matches(client, admin_header):
    response = client.get(
        "/api/v1/modules/master-data/batches/summary",
        params={"template_code": "REGIONS_BASIC", "file_name_contains": "does_not_exist"},
        headers=admin_header,
    )

    assert response.status_code == 200
    assert response.json() == {
        "total_batches": 0,
        "total_rows": 0,
        "total_issues": 0,
        "latest_batch_id": None,
        "status_breakdown": [],
        "template_breakdown": [],
    }


def test_master_data_batch_build_csv_is_idempotent_for_double_click(
    client,
    admin_header,
    db_session,
):
    workbook = Workbook()
    regions = workbook.active
    regions.title = "REGIONS"
    regions.append(["Region GID", "Region XID", "Region Name"])
    regions.append(["SYN.REGION_001", "REGION_001", "Synthetic Region"])
    details = workbook.create_sheet("REGION_DETAILS")
    details.append(["Region GID", "Location GID"])
    details.append(["SYN.REGION_001", "SYN.LOCATION_001"])
    workbook_bytes = BytesIO()
    workbook.save(workbook_bytes)
    workbook_bytes.seek(0)
    batch_response = client.post(
        "/api/v1/modules/master-data/templates/REGIONS_BASIC/batches",
        headers=admin_header,
        files={
            "file": (
                "regions_basic_double_click_upload.xlsx",
                workbook_bytes.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    batch_id = batch_response.json()["batch_id"]
    client.post(
        f"/api/v1/modules/master-data/batches/{batch_id}/validate-relationships",
        headers=admin_header,
    )
    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/map", headers=admin_header)
    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/build-output", headers=admin_header)
    first_response = client.post(
        f"/api/v1/modules/master-data/batches/{batch_id}/build-csv",
        headers=admin_header,
    )

    second_response = client.post(
        f"/api/v1/modules/master-data/batches/{batch_id}/build-csv",
        headers=admin_header,
    )

    assert first_response.status_code == 200
    assert second_response.status_code == 200
    payload = second_response.json()
    assert payload["batch_id"] == batch_id
    assert payload["status"] == "CSV_BUILT"
    assert payload["csv_file_count"] == 2
    assert [item["file_name"] for item in payload["files"]] == ["001_REGION.csv", "002_REGION_DETAIL.csv"]
    persisted_count = db_session.execute(
        text("select count(*) from master_data_csv_files where batch_id = :batch_id"),
        {"batch_id": batch_id},
    ).scalar_one()
    assert persisted_count == 2


def test_master_data_batch_export_csv_package_creates_zip_manifest_and_evidence(
    client,
    admin_header,
    db_session,
):
    workbook = Workbook()
    regions = workbook.active
    regions.title = "REGIONS"
    regions.append(["Region GID", "Region XID", "Region Name"])
    regions.append(["SYN.REGION_001", "REGION_001", "Synthetic Region"])
    details = workbook.create_sheet("REGION_DETAILS")
    details.append(["Region GID", "Location GID"])
    details.append(["SYN.REGION_001", "SYN.LOCATION_001"])
    workbook_bytes = BytesIO()
    workbook.save(workbook_bytes)
    workbook_bytes.seek(0)
    batch_response = client.post(
        "/api/v1/modules/master-data/templates/REGIONS_BASIC/batches",
        headers=admin_header,
        files={
            "file": (
                "regions_basic_upload.xlsx",
                workbook_bytes.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    batch_id = batch_response.json()["batch_id"]
    client.post(
        f"/api/v1/modules/master-data/batches/{batch_id}/validate-relationships",
        headers=admin_header,
    )
    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/map", headers=admin_header)
    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/build-output", headers=admin_header)
    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/build-csv", headers=admin_header)

    response = client.post(
        f"/api/v1/modules/master-data/batches/{batch_id}/export-csv-package",
        headers=admin_header,
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["batch_id"] == batch_id
    assert payload["status"] == "EXPORTED"
    assert payload["file_name"].startswith("master_data_batch_")
    assert payload["file_name"].endswith(".zip")
    assert payload["tables"] == ["REGION", "REGION_DETAIL"]

    artifact = db_session.query(Artifact).filter(Artifact.id == payload["artifact_id"]).one()
    manifest = db_session.query(Manifest).filter(Manifest.id == payload["manifest_id"]).one()
    evidence = db_session.query(Evidence).filter(Evidence.id == payload["evidence_id"]).one()
    audit = db_session.query(AuditLog).filter(AuditLog.action == "master_data.batch.export_csv").one()
    assert artifact.source_module == "master_data"
    assert artifact.artifact_type == "master_data_csv_zip"
    assert artifact.content_type == "application/zip"
    assert artifact.sensitivity_level == "client_safe"
    assert evidence.evidence_type == "master_data_csv_export"
    assert evidence.client_safe is True
    assert str(artifact.id) in audit.metadata_json

    with zipfile.ZipFile(artifact.file_path) as archive:
        names = sorted(archive.namelist())
        assert names == ["csv/001_REGION.csv", "csv/002_REGION_DETAIL.csv", "manifest.json"]
        csv_text = archive.read("csv/001_REGION.csv").decode("utf-8")
        manifest_payload = json.loads(archive.read("manifest.json").decode("utf-8"))

    assert csv_text.splitlines()[0] == "REGION"
    assert manifest_payload["manifest_type"] == "master_data_csv_export"
    assert manifest_payload["schema_version"] == "master-data-csv-export-manifest/v1"
    assert manifest_payload["source_entity_id"] == batch_id
    assert manifest_payload["files"][0]["file_name"] == "001_REGION.csv"
    assert json.loads(manifest.manifest_json)["manifest_type"] == "master_data_csv_export"


def test_master_data_batch_export_csv_package_is_idempotent_for_retry(
    client,
    admin_header,
    db_session,
):
    workbook = Workbook()
    regions = workbook.active
    regions.title = "REGIONS"
    regions.append(["Region GID", "Region XID", "Region Name"])
    regions.append(["SYN.REGION_001", "REGION_001", "Synthetic Region"])
    details = workbook.create_sheet("REGION_DETAILS")
    details.append(["Region GID", "Location GID"])
    details.append(["SYN.REGION_001", "SYN.LOCATION_001"])
    workbook_bytes = BytesIO()
    workbook.save(workbook_bytes)
    workbook_bytes.seek(0)
    batch_response = client.post(
        "/api/v1/modules/master-data/templates/REGIONS_BASIC/batches",
        headers=admin_header,
        files={
            "file": (
                "regions_basic_export_retry_upload.xlsx",
                workbook_bytes.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    batch_id = batch_response.json()["batch_id"]
    client.post(
        f"/api/v1/modules/master-data/batches/{batch_id}/validate-relationships",
        headers=admin_header,
    )
    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/map", headers=admin_header)
    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/build-output", headers=admin_header)
    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/build-csv", headers=admin_header)
    first_response = client.post(
        f"/api/v1/modules/master-data/batches/{batch_id}/export-csv-package",
        headers=admin_header,
    )

    second_response = client.post(
        f"/api/v1/modules/master-data/batches/{batch_id}/export-csv-package",
        headers=admin_header,
    )

    assert first_response.status_code == 200
    assert second_response.status_code == 200
    first_payload = first_response.json()
    second_payload = second_response.json()
    assert second_payload["status"] == "EXPORTED"
    assert second_payload["artifact_id"] == first_payload["artifact_id"]
    assert second_payload["manifest_id"] == first_payload["manifest_id"]
    assert second_payload["evidence_id"] == first_payload["evidence_id"]
    assert second_payload["tables"] == ["REGION", "REGION_DETAIL"]
    assert (
        db_session.query(Artifact)
        .filter(Artifact.source_module == "master_data")
        .filter(Artifact.artifact_type == "master_data_csv_zip")
        .count()
        == 1
    )
    assert (
        db_session.query(Evidence)
        .filter(Evidence.source_module == "master_data")
        .filter(Evidence.evidence_type == "master_data_csv_export")
        .count()
        == 1
    )
    assert db_session.query(AuditLog).filter(AuditLog.action == "master_data.batch.export_csv").count() == 1


def test_master_data_batch_detail_and_artifacts_are_client_safe_and_downloadable(
    client,
    admin_header,
):
    workbook = Workbook()
    regions = workbook.active
    regions.title = "REGIONS"
    regions.append(["Region GID", "Region XID", "Region Name"])
    regions.append(["SYN.REGION_001", "REGION_001", "Synthetic Region"])
    details = workbook.create_sheet("REGION_DETAILS")
    details.append(["Region GID", "Location GID"])
    details.append(["SYN.REGION_001", "SYN.LOCATION_001"])
    workbook_bytes = BytesIO()
    workbook.save(workbook_bytes)
    workbook_bytes.seek(0)
    batch_response = client.post(
        "/api/v1/modules/master-data/templates/REGIONS_BASIC/batches",
        headers=admin_header,
        files={
            "file": (
                "regions_basic_upload.xlsx",
                workbook_bytes.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    batch_id = batch_response.json()["batch_id"]
    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/validate-relationships", headers=admin_header)
    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/map", headers=admin_header)
    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/build-output", headers=admin_header)
    client.post(f"/api/v1/modules/master-data/batches/{batch_id}/build-csv", headers=admin_header)
    export_response = client.post(
        f"/api/v1/modules/master-data/batches/{batch_id}/export-csv-package",
        headers=admin_header,
    )
    artifact_id = export_response.json()["artifact_id"]

    list_response = client.get("/api/v1/modules/master-data/batches", headers=admin_header)
    detail_response = client.get(f"/api/v1/modules/master-data/batches/{batch_id}", headers=admin_header)
    artifacts_response = client.get(f"/api/v1/modules/master-data/batches/{batch_id}/artifacts", headers=admin_header)
    download_response = client.get(
        f"/api/v1/modules/master-data/batches/{batch_id}/artifacts/{artifact_id}/download",
        headers=admin_header,
    )

    assert list_response.status_code == 200
    assert list_response.json()["items"][0]["batch_id"] == batch_id
    assert "file_path" not in json.dumps(list_response.json())
    assert detail_response.status_code == 200
    detail_payload = detail_response.json()
    assert detail_payload["batch_id"] == batch_id
    assert detail_payload["status"] == "EXPORTED"
    assert detail_payload["sheet_summaries"][0]["sheet_code"] == "REGIONS"
    assert detail_payload["csv_file_count"] == 2
    assert "file_path" not in json.dumps(detail_payload)
    assert artifacts_response.status_code == 200
    artifact_payload = artifacts_response.json()["items"][0]
    assert artifact_payload["id"] == artifact_id
    assert artifact_payload["download_url"].endswith(
        f"/api/v1/modules/master-data/batches/{batch_id}/artifacts/{artifact_id}/download"
    )
    assert "file_path" not in json.dumps(artifacts_response.json())
    assert download_response.status_code == 200
    assert download_response.headers["content-type"] == "application/zip"
    assert download_response.headers["x-artifact-sha256"] == export_response.json()["sha256"]
    assert download_response.content.startswith(b"PK")


def test_master_data_batch_artifact_download_rejects_cross_batch_artifact(
    client,
    admin_header,
):
    first_workbook = Workbook()
    first_regions = first_workbook.active
    first_regions.title = "REGIONS"
    first_regions.append(["Region GID", "Region XID", "Region Name"])
    first_regions.append(["SYN.REGION_001", "REGION_001", "Synthetic Region"])
    first_details = first_workbook.create_sheet("REGION_DETAILS")
    first_details.append(["Region GID", "Location GID"])
    first_details.append(["SYN.REGION_001", "SYN.LOCATION_001"])
    first_bytes = BytesIO()
    first_workbook.save(first_bytes)
    first_bytes.seek(0)
    first_batch = client.post(
        "/api/v1/modules/master-data/templates/REGIONS_BASIC/batches",
        headers=admin_header,
        files={
            "file": (
                "regions_basic_upload_1.xlsx",
                first_bytes.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    ).json()["batch_id"]
    client.post(f"/api/v1/modules/master-data/batches/{first_batch}/validate-relationships", headers=admin_header)
    client.post(f"/api/v1/modules/master-data/batches/{first_batch}/map", headers=admin_header)
    client.post(f"/api/v1/modules/master-data/batches/{first_batch}/build-output", headers=admin_header)
    client.post(f"/api/v1/modules/master-data/batches/{first_batch}/build-csv", headers=admin_header)
    export = client.post(
        f"/api/v1/modules/master-data/batches/{first_batch}/export-csv-package",
        headers=admin_header,
    ).json()

    second_workbook = Workbook()
    second_regions = second_workbook.active
    second_regions.title = "REGIONS"
    second_regions.append(["Region GID", "Region XID", "Region Name"])
    second_regions.append(["SYN.REGION_002", "REGION_002", "Synthetic Region 2"])
    second_details = second_workbook.create_sheet("REGION_DETAILS")
    second_details.append(["Region GID", "Location GID"])
    second_details.append(["SYN.REGION_002", "SYN.LOCATION_002"])
    second_bytes = BytesIO()
    second_workbook.save(second_bytes)
    second_bytes.seek(0)
    second_batch = client.post(
        "/api/v1/modules/master-data/templates/REGIONS_BASIC/batches",
        headers=admin_header,
        files={
            "file": (
                "regions_basic_upload_2.xlsx",
                second_bytes.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    ).json()["batch_id"]

    response = client.get(
        f"/api/v1/modules/master-data/batches/{second_batch}/artifacts/{export['artifact_id']}/download",
        headers=admin_header,
    )

    assert response.status_code == 404
    assert response.json()["code"] == "MASTER_DATA_ARTIFACT_NOT_FOUND"


def test_master_data_batch_artifacts_marks_missing_file_unavailable(
    client,
    admin_header,
    db_session,
):
    workbook = Workbook()
    regions = workbook.active
    regions.title = "REGIONS"
    regions.append(["Region GID", "Region XID", "Region Name"])
    regions.append(["SYN.REGION_001", "REGION_001", "Synthetic Region"])
    details = workbook.create_sheet("REGION_DETAILS")
    details.append(["Region GID", "Location GID"])
    details.append(["SYN.REGION_001", "SYN.LOCATION_001"])
    workbook_bytes = BytesIO()
    workbook.save(workbook_bytes)
    workbook_bytes.seek(0)
    batch_id = client.post(
        "/api/v1/modules/master-data/templates/REGIONS_BASIC/batches",
        headers=admin_header,
        files={
            "file": (
                "regions_basic_missing_artifact_upload.xlsx",
                workbook_bytes.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    ).json()["batch_id"]
    missing_path = Path("var/artifacts/master_data/removed_export.zip")
    artifact = Artifact(
        source_module="master_data",
        artifact_type="master_data_csv_zip",
        file_path=str(missing_path),
        file_name="removed_export.zip",
        content_type="application/zip",
        sha256="0" * 64,
        size_bytes=123,
        sensitivity_level="client_safe",
    )
    db_session.add(artifact)
    db_session.flush()
    db_session.add(
        Evidence(
            source_module="master_data",
            evidence_type="master_data_csv_export",
            status="CREATED",
            summary_json=json.dumps({"source_entity_id": batch_id}),
            artifact_id=artifact.id,
            client_safe=True,
            sensitivity_level="client_safe",
        )
    )
    db_session.commit()

    response = client.get(
        f"/api/v1/modules/master-data/batches/{batch_id}/artifacts",
        headers=admin_header,
    )

    assert response.status_code == 200
    item = response.json()["items"][0]
    assert item["id"] == artifact.id
    assert item["availability_status"] == "FILE_MISSING"
    assert item["download_url"] is None
